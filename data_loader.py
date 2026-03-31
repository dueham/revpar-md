"""
data_loader.py  —  RevPar MD Platform (AWS Version)
=====================================================
All data loading and parsing logic. Each function reads one source file
and returns clean pandas DataFrames ready for the dashboard.

AWS NOTE: All file paths are S3 keys (strings like "Hampton_Superior/STR.xlsx").
The _open() helper transparently fetches from S3 and returns a BytesIO object,
which both openpyxl and pandas accept natively in place of a file path.

Caching: Streamlit's @st.cache_data(ttl=300) is applied at the call site
in the app, so this module stays framework-agnostic and testable.
"""

import io
import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime, date
from typing import Optional
import warnings
warnings.filterwarnings("ignore")

from hotel_config import download_file_bytes


def _open(path) -> io.BytesIO:
    """
    Universal file opener for AWS.
    - If path is already a BytesIO, return it as-is.
    - If path is an S3 key string (e.g. 'Hampton_Superior/STR.xlsx'), fetch from S3.
    - If path is a local Path or file string (fallback for testing), open locally.
    """
    if isinstance(path, io.BytesIO):
        path.seek(0)
        return path
    if isinstance(path, str):
        # S3 key — fetch from bucket
        buf = download_file_bytes(path)
        if buf is not None:
            return buf
        # Fallback: try as local path (useful for local dev/testing)
        return io.BytesIO(open(path, "rb").read())
    if isinstance(path, Path):
        return io.BytesIO(path.read_bytes())
    raise ValueError(f"Unsupported path type: {type(path)}")


# ═══════════════════════════════════════════════════════════════════════════════
# YEAR.XLSX  —  Master demand/forecast/actuals extract
# 762 days · Jan 2025–Feb 2027 · 70 columns
# ═══════════════════════════════════════════════════════════════════════════════

def load_year(path) -> pd.DataFrame:
    """
    Reads year.xlsx (Property sheet, header row 6, data from row 7).
    Returns a clean DataFrame with parsed dates and key columns renamed.
    """
    df = pd.read_excel(_open(path), sheet_name="Property", header=5, engine="openpyxl")
    df.columns = df.columns.str.strip()

    # Parse date
    df["Date"] = pd.to_datetime(df["Occupancy Date"], errors="coerce")
    df = df.dropna(subset=["Date"])
    df = df.sort_values("Date").reset_index(drop=True)

    # Rename most-used columns to short names.
    # Two naming conventions exist across hotels:
    #   "Last Year Actual" — standard Hilton Concerto export
    #   "STLY"             — same-time-last-year variant (e.g. Cherry Creek export)
    # Both are mapped to the same internal short names so all tabs work uniformly.
    rename = {
        "Day of Week":                                   "DayOfWeek",
        "Special Event This Year":                       "Event",
        "Physical Capacity This Year":                   "Capacity",
        "Occupancy On Books This Year":                  "OTB",
        # LY / STLY variants
        "Occupancy On Books Last Year Actual":           "OTB_LY",
        "Occupancy On Books STLY":                       "OTB_LY",
        "Rooms Sold - Group This Year":                  "Group_OTB",
        "Rooms Sold - Group STLY":                       "Group_OTB_LY",
        "Rooms Sold - Transient This Year":              "Transient_OTB",
        "Rooms Sold - Transient STLY":                   "Transient_OTB_LY",
        "Booked Room Revenue This Year":                 "Revenue_OTB",
        "Booked Room Revenue Last Year Actual":          "Revenue_LY",
        "Booked Room Revenue STLY":                      "Revenue_LY",
        "Forecasted Room Revenue This Year":             "Revenue_Forecast",
        "Occupancy Forecast - Total This Year":          "Forecast_Rooms",
        "Occupancy Forecast - Total Last Year Actual":   "Forecast_Rooms_LY",
        "Occupancy Forecast - Total STLY":               "Forecast_Rooms_LY",
        "RevPAR On Books This Year":                     "RevPAR_OTB",
        "RevPAR Forecast This Year":                     "RevPAR_Forecast",
        "ADR On Books This Year":                        "ADR_OTB",
        "ADR Forecast This Year":                        "ADR_Forecast",
        "Rooms N/A - Out of Order This Year":            "OOO",
        "Wash % This Year":                              "Wash_Pct",
        "LV0":                                           "LV0_Rate",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})

    # If both "Last Year Actual" and "STLY" variants existed, dedupe — keep first occurrence
    for short in ["OTB_LY", "Revenue_LY", "Forecast_Rooms_LY"]:
        dupes = [c for c in df.columns if c == short]
        if len(dupes) > 1:
            df = df.loc[:, ~df.columns.duplicated(keep="first")]

    # Derived fields
    df["Occ_Pct_OTB"] = (df["OTB"] / df["Capacity"] * 100).round(1)

    return df


def get_tonight_otb(year_df: pd.DataFrame, total_rooms: int) -> dict:
    """Extract tonight's OTB stats for the landing page card."""
    today = pd.Timestamp(datetime.now().date())
    row = year_df[year_df["Date"] == today]
    if row.empty:
        # Try tomorrow if after midnight cutoff
        row = year_df[year_df["Date"] == today + pd.Timedelta(days=1)]
    if row.empty:
        return {"occ_pct": None, "otb_rooms": None, "adr": None, "event": None}
    r = row.iloc[0]
    otb = r.get("OTB", 0) or 0
    cap = r.get("Capacity", total_rooms) or total_rooms
    return {
        "occ_pct":   round(otb / cap * 100, 1) if cap else None,
        "otb_rooms": int(otb),
        "adr":       round(r.get("ADR_OTB", 0), 2) if r.get("ADR_OTB") else None,
        "event":     r.get("Event", "") or "",
        "date":      today,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# BUDGET.XLSX  —  Full year 2026 daily budget
# 366 days · Rooms, Occ%, ADR, RevPAR, Revenue
# ═══════════════════════════════════════════════════════════════════════════════

def load_budget(path) -> pd.DataFrame:
    """
    Reads Budget.xlsx (header row 3, data from row 4).
    Layout: Date | Rooms | Occ Rooms | Occ% | ADR | RevPAR | Revenue
    Auto-detects sheet name — tries Sheet1, Budget, then falls back to
    the first sheet, so this works regardless of how the tab is named.
    """
    wb = openpyxl.load_workbook(_open(path), data_only=True)
    preferred = ["Sheet1", "Budget"]
    sheet_name = next((n for n in preferred if n in wb.sheetnames), wb.sheetnames[0])

    df = pd.read_excel(_open(path), sheet_name=sheet_name, header=2, engine="openpyxl")
    # Trim to exactly 7 columns — IHG budget files may have extra trailing columns
    df = df.iloc[:, :7]
    df.columns = ["Date_Raw", "Rooms", "Occ_Rooms", "Occ_Pct", "ADR", "RevPAR", "Revenue"]
    df["Date"] = pd.to_datetime(df["Date_Raw"], errors="coerce")
    df = df.dropna(subset=["Date"])
    df = df[df["Date_Raw"] != "Total"].reset_index(drop=True)
    df = df.sort_values("Date").reset_index(drop=True)
    for col in ["Rooms", "Occ_Rooms", "Occ_Pct", "ADR", "RevPAR", "Revenue"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df[["Date", "Rooms", "Occ_Rooms", "Occ_Pct", "ADR", "RevPAR", "Revenue"]]


# ═══════════════════════════════════════════════════════════════════════════════
# 1.XLSX / 7.XLSX  —  Change & Differential Control Reports
# Overnight pickup (1.xlsx) and 7-day pickup (7.xlsx)
# ═══════════════════════════════════════════════════════════════════════════════

def load_pickup(path) -> pd.DataFrame:
    """
    Reads 1.xlsx or 7.xlsx.
    Header is rows 6–7 (merged), data starts row 8.
    Columns: DayOfWeek, Date, Event, OOO, OTB_Current, OTB_Change,
             Forecast_Current, Forecast_Change
    """
    # NOTE: Must NOT use read_only=True — Hilton-exported XLSX files return only 1 row in read_only mode
    wb = openpyxl.load_workbook(_open(path), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Auto-detect first data row: find first row where col[1] looks like a date
    # (handles files where header rows differ from expected row 8)
    start_idx = 7  # default: row 8 (index 7)
    for i, row in enumerate(rows):
        if row is None or len(row) < 2:
            continue
        try:
            val = pd.to_datetime(row[1], errors="coerce")
            if val is not pd.NaT and not pd.isna(val) and i >= 4:
                start_idx = i
                break
        except Exception:
            continue

    records = []
    for row in rows[start_idx:]:
        if row is None or len(row) < 10:
            continue
        dt = pd.to_datetime(row[1], errors="coerce")
        if dt is pd.NaT or pd.isna(dt):
            continue
        records.append({
            "DayOfWeek":        row[0],
            "Date":             dt,
            "Event":            row[3] or "",
            "OOO":              row[5],
            "OTB_Current":      row[7],
            "OTB_Change":       row[9],
            "Forecast_Current": row[12] if len(row) > 12 else None,
            "Forecast_Change":  row[14] if len(row) > 14 else None,
        })

    df = pd.DataFrame(records)
    if df.empty:
        return df
    df["Date"] = pd.to_datetime(df["Date"]).dt.normalize()
    df = df.dropna(subset=["Date"]).sort_values("Date").reset_index(drop=True)
    return df


# ═══════════════════════════════════════════════════════════════════════════════
# GROUP_WASH_REPORT.XLSX  —  Group block/pickup by night
# 195 rows · Block, Pickup, Wash, Rate, Days to Arrival, Sales Manager
# ═══════════════════════════════════════════════════════════════════════════════

def load_groups(path) -> pd.DataFrame:
    """
    Reads Group_Wash_Report.xlsx (Report sheet, header row 2).
    Cleans column names and converts numeric fields.
    """
    df = pd.read_excel(_open(path), sheet_name="Report", header=1, engine="openpyxl")

    # Normalize column names (remove newlines)
    df.columns = df.columns.str.replace("\n", " ").str.strip()

    rename = {
        "Group Code":           "Group_Code",
        "Group Name":           "Group_Name",
        "Market Segment":       "Segment",
        "Forecast Group":       "Forecast_Type",
        "From":                 "Arrival",
        "To":                   "Departure",
        "Sales Manager":        "Sales_Manager",
        "Cutoff Date":          "Cutoff",
        "Day of Week":          "DayOfWeek",
        "Occupancy Date":       "Occ_Date",
        "Group Arrivals":       "Arrivals",
        "Block":                "Block",
        "Pick Up":              "Pickup",
        "Available Block":      "Avail_Block",
        "7 Day Pickup Variance":"Pickup_7Day_Var",
        "Rate":                 "Rate",
        "System Wash %":        "System_Wash_Pct",
        "User Wash %":          "User_Wash_Pct",
        "Days to Arrival":      "Days_To_Arrival",
        "Notes":                "Notes",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})

    # Parse dates
    for col in ["Arrival", "Departure", "Cutoff", "Occ_Date"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    # Numeric conversions
    for col in ["Block", "Pickup", "Avail_Block", "Rate", "System_Wash_Pct",
                "User_Wash_Pct", "Days_To_Arrival", "Pickup_7Day_Var"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Derived
    if "Block" in df.columns and "Pickup" in df.columns:
        df["Pickup_Pct"] = (df["Pickup"] / df["Block"] * 100).where(df["Block"] > 0).round(1)

    df = df.dropna(subset=["Group_Name"] if "Group_Name" in df.columns else [])
    return df


# ═══════════════════════════════════════════════════════════════════════════════
# CORP_SEGMENTS.XLSX  —  IHG Corporate Account & Segment Production Report
# ═══════════════════════════════════════════════════════════════════════════════

def load_ihg_corp_segments(path) -> dict:
    """
    Reads IHG Corp_Segments.xlsx.
    Returns dict with two DataFrames:
      'segments'  — one row per Segment, summed CY/LY rooms/revenue/ADR
      'companies' — one row per Company Name, summed, dominant segment shown
    Rows where both CY_Rooms and LY_Rooms are zero are excluded.
    Blank Company Name rows are excluded.
    ADR is derived (Rev/Rooms) for weighted accuracy.
    """
    wb = openpyxl.load_workbook(_open(path), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"companies": pd.DataFrame(), "segments": pd.DataFrame()}

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_map  = {h: i for i, h in enumerate(headers)}

    ci_name   = col_map.get("Company Name")
    ci_seg    = col_map.get("Segment")
    ci_cy_rms = col_map.get("CY Rooms Qty")
    ci_ly_rms = col_map.get("LY Rooms Qty")
    ci_cy_rev = col_map.get("CY Revenue")
    ci_ly_rev = col_map.get("LY Revenue")

    if any(c is None for c in [ci_name, ci_seg, ci_cy_rms, ci_ly_rms, ci_cy_rev, ci_ly_rev]):
        return {"companies": pd.DataFrame(), "segments": pd.DataFrame()}

    records = []
    for row in rows[1:]:
        if row is None: continue
        name = str(row[ci_name]).strip() if row[ci_name] else ""
        seg  = str(row[ci_seg]).strip()  if row[ci_seg]  else ""
        if not name: continue
        def _n(v):
            try: return float(v) if v is not None else 0.0
            except: return 0.0
        cy_r = _n(row[ci_cy_rms]); ly_r = _n(row[ci_ly_rms])
        cy_v = _n(row[ci_cy_rev]); ly_v = _n(row[ci_ly_rev])
        if cy_r == 0 and ly_r == 0: continue
        records.append({"CompanyName": name, "Segment": seg,
                        "CY_Rooms": cy_r, "LY_Rooms": ly_r,
                        "CY_Rev": cy_v,   "LY_Rev": ly_v})

    if not records:
        return {"companies": pd.DataFrame(), "segments": pd.DataFrame()}

    df = pd.DataFrame(records)

    # Segment rollup
    seg_grp = df.groupby("Segment", as_index=False).agg(
        CY_Rooms=("CY_Rooms","sum"), LY_Rooms=("LY_Rooms","sum"),
        CY_Rev=("CY_Rev","sum"),     LY_Rev=("LY_Rev","sum"))
    seg_grp["CY_ADR"]    = (seg_grp["CY_Rev"]/seg_grp["CY_Rooms"]).where(seg_grp["CY_Rooms"]>0).round(2)
    seg_grp["LY_ADR"]    = (seg_grp["LY_Rev"]/seg_grp["LY_Rooms"]).where(seg_grp["LY_Rooms"]>0).round(2)
    seg_grp["Var_Rooms"] = seg_grp["CY_Rooms"] - seg_grp["LY_Rooms"]
    seg_grp["Var_Rev"]   = (seg_grp["CY_Rev"]  - seg_grp["LY_Rev"]).round(2)
    seg_grp["Var_ADR"]   = (seg_grp["CY_ADR"].fillna(0) - seg_grp["LY_ADR"].fillna(0)).round(2)
    seg_grp = seg_grp.sort_values("CY_Rooms", ascending=False).reset_index(drop=True)

    # Company rollup — dominant segment = highest CY_Rooms segment per company
    dom = (df[df["CY_Rooms"]>0]
           .groupby(["CompanyName","Segment"])["CY_Rooms"].sum().reset_index()
           .sort_values("CY_Rooms", ascending=False)
           .drop_duplicates("CompanyName")
           .rename(columns={"Segment":"Dom_Segment"})[["CompanyName","Dom_Segment"]])
    co_grp = df.groupby("CompanyName", as_index=False).agg(
        CY_Rooms=("CY_Rooms","sum"), LY_Rooms=("LY_Rooms","sum"),
        CY_Rev=("CY_Rev","sum"),     LY_Rev=("LY_Rev","sum"))
    co_grp = co_grp.merge(dom, on="CompanyName", how="left")
    co_grp["Dom_Segment"].fillna("", inplace=True)
    co_grp["CY_ADR"]    = (co_grp["CY_Rev"]/co_grp["CY_Rooms"]).where(co_grp["CY_Rooms"]>0).round(2)
    co_grp["LY_ADR"]    = (co_grp["LY_Rev"]/co_grp["LY_Rooms"]).where(co_grp["LY_Rooms"]>0).round(2)
    co_grp["Var_Rooms"] = co_grp["CY_Rooms"] - co_grp["LY_Rooms"]
    co_grp["Var_Rev"]   = (co_grp["CY_Rev"]  - co_grp["LY_Rev"]).round(2)
    co_grp["Var_ADR"]   = (co_grp["CY_ADR"].fillna(0) - co_grp["LY_ADR"].fillna(0)).round(2)
    co_grp = co_grp.sort_values("CY_Rooms", ascending=False).reset_index(drop=True)

    return {"companies": co_grp, "segments": seg_grp}



# ═══════════════════════════════════════════════════════════════════════════════
# IHG GROUP SUMMARY BY DAY  —  Groups.xlsx export from IHG system
# ═══════════════════════════════════════════════════════════════════════════════

def load_ihg_groups(path) -> pd.DataFrame:
    """
    Reads IHG 'Group Summary by Day' Excel export.
    Two-level structure:
      Lvl1 rows — group header: Code, Name, Arrival, Departure, Release Date, Status, Segment
      Lvl2 rows — nightly detail: Stay-over Date, CT (block), RS (pickup), AU (avail), Rate

    Confirmed column layout (0-based) from actual file header:
      [0]=Level, [1]=Code, [2]=Name, [3]=Arrival, [4]=Departure,
      [5]=Housing Status, [6]=Rolling, [7]=Release Date, [8]=Release Days,
      [9]=Market Segmentation, [10]=Industry Type,
      [11]=Lvl2-Stay-over Date, [12]=Lvl2-CT (block), [13]=Lvl2-RS (pickup),
      [14]=Lvl2-AU (avail — always equals CT, not remaining), [15]=Lvl2-Lowest Rate

    Note: AU [14] always equals CT [12] in IHG exports — true availability is
    calculated as Block - Pickup after aggregation, not from AU directly.

    Only groups without 'Cancelled' in name/status are returned.
    Returns same schema as load_groups() so render_groups_tab works unchanged.
    """
    wb = openpyxl.load_workbook(_open(path), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (col A = 'Level')
    header_idx = None
    for i, row in enumerate(rows):
        if row and str(row[0]).strip() == 'Level':
            header_idx = i
            break
    if header_idx is None:
        return pd.DataFrame()

    def _parse_ihg_date(val):
        if val is None: return pd.NaT
        s = str(val).strip().split()[0]
        try: return pd.to_datetime(s, format='%d%b%Y')
        except: return pd.NaT

    def _safe_int(val):
        """Parse integer safely, return 0 for None/non-numeric values."""
        if val is None: return 0
        try: return int(float(str(val).strip()))
        except: return 0

    def _safe_rate(val):
        if val is None or str(val).strip() in ('-', '', 'False', 'True'): return None
        try: return float(val)
        except: return None

    today = pd.Timestamp(datetime.now().date())
    records = []
    current_group = {}

    for row in rows[header_idx + 1:]:
        if row is None or row[0] is None: continue
        level = str(row[0]).strip()

        if level == 'Lvl1':
            name_val = str(row[2]).strip() if row[2] else ''
            # Skip cancelled groups — check both name and [9] status field if present
            status_9 = str(row[9]).strip().lower() if row[9] else ''
            if not name_val or 'cancelled' in name_val.lower() or status_9 == 'cancelled':
                current_group = {}
                continue
            current_group = {
                'Group_Code': row[1],
                'Group_Name': row[2],
                'Arrival':    _parse_ihg_date(row[3]),
                'Departure':  _parse_ihg_date(row[4]),
                'Cutoff':     _parse_ihg_date(row[7]),
                'Status':     str(row[5]).strip() if row[5] else '',
                'Segment':    str(row[9]).strip() if row[9] else '',
            }

        elif level == 'Lvl2' and current_group:
            # [11]=Stay-over Date, [12]=CT(block), [13]=RS(pickup),
            # [14]=AU (equals CT, not remaining avail), [15]=Lowest Rate
            occ_date = _parse_ihg_date(row[11])
            block    = _safe_int(row[12])
            pickup   = _safe_int(row[13])
            # Avail_Block = Block - Pickup (AU col always equals CT, not remaining)
            avail    = max(0, block - pickup)
            rate     = _safe_rate(row[15])
            days_to_arrival = int((current_group['Arrival'] - today).days) \
                if pd.notna(current_group['Arrival']) else None
            pickup_pct = round(pickup / block * 100, 1) if block > 0 else 0.0
            records.append({
                **current_group,
                'Occ_Date':        occ_date,
                'Block':           block,
                'Pickup':          pickup,
                'Avail_Block':     avail,
                'Rate':            rate,
                'Pickup_Pct':      pickup_pct,
                'Days_To_Arrival': days_to_arrival,
                'Sales_Manager':   None,
                'Pickup_7Day_Var': None,
            })

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)
    for col in ['Arrival', 'Departure', 'Cutoff', 'Occ_Date']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    return df.sort_values(['Arrival', 'Group_Name', 'Occ_Date']).reset_index(drop=True)


# ═══════════════════════════════════════════════════════════════════════════════
# BOOKING_REPORTS_SRP_ACTIVITY.XLSX  —  Individual reservation records
# 24,623 records · Jan 2024–Sep 2026 · 12 MCAT segments
# ═══════════════════════════════════════════════════════════════════════════════

def load_srp(path) -> pd.DataFrame:
    """
    Reads Booking_Reports_SRP_Activity.xlsx (SRP Activity sheet).
    Header is row 13 (index 12). Returns clean reservation-level DataFrame.
    """
    df = pd.read_excel(
        path,
        sheet_name="SRP Activity",
        header=12,
        engine="openpyxl",
        usecols=[0, 1, 2, 3, 4, 8, 9, 10, 11, 12, 14, 15]
    )

    col_map = {
        df.columns[0]:  "Stay_ID",
        df.columns[1]:  "Confirmation",
        df.columns[2]:  "Booked_Date",
        df.columns[3]:  "Arrival_Date",
        df.columns[4]:  "Departure_Date",
        df.columns[5]:  "Property",
        df.columns[6]:  "SRP_Code",
        df.columns[7]:  "SRP_Type",
        df.columns[8]:  "SRP_Name",
        df.columns[9]:  "MCAT",
        df.columns[10]: "Room_Nights",
        df.columns[11]: "Room_Revenue",
    }
    df = df.rename(columns=col_map)

    for col in ["Booked_Date", "Arrival_Date", "Departure_Date"]:
        df[col] = pd.to_datetime(df[col], errors="coerce")

    df["Room_Nights"]  = pd.to_numeric(df["Room_Nights"],  errors="coerce")
    df["Room_Revenue"] = pd.to_numeric(df["Room_Revenue"], errors="coerce")
    df["ADR"]          = (df["Room_Revenue"] / df["Room_Nights"]).where(df["Room_Nights"] > 0).round(2)

    df = df.dropna(subset=["Arrival_Date"]).reset_index(drop=True)
    return df


# ═══════════════════════════════════════════════════════════════════════════════
# RATES.XLSX  —  Live BAR rate shopping vs comp set
# 5 sheets · 365 days forward · 6 comp hotels
# ═══════════════════════════════════════════════════════════════════════════════

# ── PER-HOTEL COMP SETS ───────────────────────────────────────────────────────
# Each hotel_id maps to its own COMP_HOTELS list and COMP_SHORT dict.
# The rate-loading functions use get_comp_set(hotel_id) to pick the right one.

_COMP_REGISTRY = {

    "hampton_lino_lakes": {
        "hotels": [
            "Hampton Inn & Suites Lino Lakes",
            "Country Inn & Suites by Radisson, Shoreview-Mounds View, MN",
            "Best Western Plus Blaine at the National Sports Center",
            "Best Western Plus White Bear Country Inn",
            "Holiday Inn Express Hotel & Suites Coon Rapids - Blaine Area by IHG",
            "AmericInn by Wyndham Mounds View Minneapolis",
            "Fairfield Inn & Suites by Marriott Minneapolis North/Blaine",
        ],
        "short": {
            "Hampton Inn & Suites Lino Lakes":                                     "Hampton (Us)",
            "Country Inn & Suites by Radisson, Shoreview-Mounds View, MN":        "Country Inn",
            "Best Western Plus Blaine at the National Sports Center":              "BW Blaine",
            "Best Western Plus White Bear Country Inn":                            "BW White Bear",
            "Holiday Inn Express Hotel & Suites Coon Rapids - Blaine Area by IHG":"HIX Coon Rapids",
            "AmericInn by Wyndham Mounds View Minneapolis":                        "AmericInn",
            "Fairfield Inn & Suites by Marriott Minneapolis North/Blaine":         "Fairfield",
        },
    },

    "hampton_superior": {
        "hotels": [
            "Hampton Inn Superior Duluth, Wi",
            "Holiday Inn Express & Suites Superior by IHG",
            "Best Western Bridgeview Hotel",
            "Comfort Inn Duluth West",
            "Barkers Island Inn Resort & Conference Center",
            "Hampton Inn & Suites Duluth North Mn",
            "Holiday Inn & Suites Duluth-Downtown by IHG",
            "Cobblestone Hotel & Suites - Superior Duluth",
            "Radisson Hotel Duluth-Harborview",
            "Tru By Hilton Duluth Mall Area",
        ],
        "short": {
            "Hampton Inn Superior Duluth, Wi":                   "Hampton (Us)",
            "Holiday Inn Express & Suites Superior by IHG":      "HIX Superior",
            "Best Western Bridgeview Hotel":                     "BW Bridgeview",
            "Comfort Inn Duluth West":                           "Comfort Inn",
            "Barkers Island Inn Resort & Conference Center":     "Barkers Island",
            "Hampton Inn & Suites Duluth North Mn":              "Hampton Duluth N",
            "Holiday Inn & Suites Duluth-Downtown by IHG":       "HIX Duluth",
            "Cobblestone Hotel & Suites - Superior Duluth":      "Cobblestone",
            "Radisson Hotel Duluth-Harborview":                  "Radisson",
            "Tru By Hilton Duluth Mall Area":                    "Tru Hilton",
        },
    },

    "hilton_checkers_la": {
        "hotels": [
            "Sheraton Grand Los Angeles",
            "Courtyard by Marriott Los Angeles L.A. LIVE",
            "Hotel Indigo Los Angeles Downtown",
            "Hilton Checkers Los Angeles",
            "The Wayfarer Downtown LA, Tapestry Collection by Hilton",
            "Hotel Figueroa, Unbound Collection by Hyatt",
            "The O Hotel, Trademark Collection by Wyndham",
            "The Delphi Downtown LA",
        ],
        "short": {
            "Sheraton Grand Los Angeles":                               "Sheraton Grand",
            "Courtyard by Marriott Los Angeles L.A. LIVE":             "Courtyard LALIVE",
            "Hotel Indigo Los Angeles Downtown":                        "Hotel Indigo",
            "Hilton Checkers Los Angeles":                              "Checkers (Us)",
            "The Wayfarer Downtown LA, Tapestry Collection by Hilton":  "Wayfarer",
            "Hotel Figueroa, Unbound Collection by Hyatt":              "Figueroa",
            "The O Hotel, Trademark Collection by Wyndham":             "The O Hotel",
            "The Delphi Downtown LA":                                   "Delphi",
        },
    },

    "hampton_cherry_creek": {
        "hotels": [
            "Hampton Inn & Suites Denver-Cherry Creek",
            "Hilton Garden Inn Denver/Cherry Creek",
            "DoubleTree by Hilton Denver Cherry Creek, CO",
            "Hyatt Place Denver Cherry Creek",
            "Courtyard by Marriott Denver Cherry Creek",
            "Hampton Inn & Suites Denver Tech Center",
            "Super 8 by Wyndham Aurora East",
        ],
        "short": {
            "Hampton Inn & Suites Denver-Cherry Creek":       "Hampton (Us)",
            "Hilton Garden Inn Denver/Cherry Creek":          "HGI Cherry Creek",
            "DoubleTree by Hilton Denver Cherry Creek, CO":   "DoubleTree CC",
            "Hyatt Place Denver Cherry Creek":                "Hyatt Place CC",
            "Courtyard by Marriott Denver Cherry Creek":      "Courtyard CC",
            "Hampton Inn & Suites Denver Tech Center":        "Hampton Tech Ctr",
            "Super 8 by Wyndham Aurora East":                 "Super 8 Aurora",
        },
    },

    "hotel_indigo_rochester": {
        # Names match exactly as they appear in Rates.xlsx row 5 header
        "hotels": [
            "Hotel Indigo Rochester - Mayo Clinic Area by IHG",
            "Hilton Garden Inn Rochester Downtown",
            "Rochester Marriott Mayo Clinic Area",
            "Hyatt House Rochester Mayo Clinic Area",
            "Hampton Inn Rochester",
            "Courtyard Rochester Mayo Clinic Area/Saint Marys",
            "Hilton Rochester Mayo Clinic Area",
            "EVEN Hotel Rochester - Mayo Clinic Area by IHG",
            "Staybridge Suites Rochester - Mayo Clinic Area by IHG",
            "Kahler Grand Hotel",
        ],
        "short": {
            "Hotel Indigo Rochester - Mayo Clinic Area by IHG":        "Hotel Indigo (Us)",
            "Hilton Garden Inn Rochester Downtown":                     "HGI Rochester",
            "Rochester Marriott Mayo Clinic Area":                      "Marriott Rochester",
            "Hyatt House Rochester Mayo Clinic Area":                   "Hyatt House",
            "Hampton Inn Rochester":                                    "Hampton Rochester",
            "Courtyard Rochester Mayo Clinic Area/Saint Marys":        "Courtyard Rochester",
            "Hilton Rochester Mayo Clinic Area":                        "Hilton Rochester",
            "EVEN Hotel Rochester - Mayo Clinic Area by IHG":          "EVEN Rochester",
            "Staybridge Suites Rochester - Mayo Clinic Area by IHG":   "Staybridge Rochester",
            "Kahler Grand Hotel":                                       "Kahler Grand",
        },
    },

    "holiday_inn_express_superior": {
        # Our hotel is not in its own IHG comp list, so we add a virtual entry with (Us) marker.
        # Comp hotel names must match exactly as they appear in StrategicAnalysis.xlsx cols 72-78.
        "hotels": [
            "Holiday Inn Express & Suites Superior (Us)",
            "Americinn By Wyndham Duluth",
            "Best Western Bridgeview Motor Superior",
            "Comfort Inn West Duluth",
            "Hampton Inn Duluth",
            "Hampton Inn Superior Duluth Wi",
            "Holiday Inn Express & Suites Duluth North - Miller Hill",
            "La Quinta Inn Suites Duluth",
        ],
        "short": {
            "Holiday Inn Express & Suites Superior (Us)":             "HIX Superior (Us)",
            "Americinn By Wyndham Duluth":                            "AmericInn",
            "Best Western Bridgeview Motor Superior":                  "BW Bridgeview",
            "Comfort Inn West Duluth":                                 "Comfort Inn",
            "Hampton Inn Duluth":                                      "Hampton Duluth",
            "Hampton Inn Superior Duluth Wi":                          "Hampton Superior",
            "Holiday Inn Express & Suites Duluth North - Miller Hill": "HIX Duluth N",
            "La Quinta Inn Suites Duluth":                             "La Quinta",
        },
    },

}

# Fallback empty set for hotels not yet configured
_EMPTY_COMP = {"hotels": [], "short": {}}


def get_comp_set(hotel_id: str) -> tuple:
    """Return (COMP_HOTELS list, COMP_SHORT dict) for the given hotel_id."""
    entry = _COMP_REGISTRY.get(hotel_id, _EMPTY_COMP)
    return entry["hotels"], entry["short"]


# ── Legacy globals (kept for backward compatibility with any direct references) ──
COMP_HOTELS = _COMP_REGISTRY["hampton_lino_lakes"]["hotels"]
COMP_SHORT  = _COMP_REGISTRY["hampton_lino_lakes"]["short"]


def load_rates_overview(path, hotel_id: str = "hampton_lino_lakes") -> pd.DataFrame:
    """
    Reads Rates.xlsx and returns: Date, Day, Our_Rate, Median_Comp, Rank, Market_Demand

    Supports two file formats:
      A) Multi-sheet (IDeaS/other): uses the "Overview" sheet, columns at fixed offsets.
      B) Lighthouse single-sheet: only a "Rates" sheet exists. Our hotel is the first
         hotel column (col D, idx 3); market demand is col C (idx 2). Median comp and
         rank are computed on-the-fly from the other hotel columns.
    """
    wb   = openpyxl.load_workbook(_open(path), read_only=True, data_only=True)
    comp_hotels, comp_short = get_comp_set(hotel_id)

    # ── Format A: dedicated Overview sheet ───────────────────────────────────
    if "Overview" in wb.sheetnames:
        ws   = wb["Overview"]
        rows = list(ws.iter_rows(values_only=True))
        records = []
        for row in rows[5:]:
            if row[2] is None:
                continue
            records.append({
                "Day":           row[1],
                "Date":          pd.to_datetime(row[2], errors="coerce"),
                "Our_Rate":      row[3],
                "Median_Comp":   row[4],
                "Rank":          row[5],
                "Market_Demand": round(row[6] * 100, 1) if row[6] else None,
                "Holiday":       row[7],
                "Event":         row[8],
            })
        df = pd.DataFrame(records).dropna(subset=["Date"]).sort_values("Date").reset_index(drop=True)
        df["Date"]        = pd.to_datetime(df["Date"]).dt.normalize()
        df["Our_Rate"]    = pd.to_numeric(df["Our_Rate"],    errors="coerce")
        df["Median_Comp"] = pd.to_numeric(df["Median_Comp"], errors="coerce")
        df["Rank"]        = pd.to_numeric(df["Rank"],        errors="coerce")
        return df

    # ── Format B: Lighthouse single Rates sheet ───────────────────────────────
    # Layout: col B = Day/Date, col C = Market Demand, col D = Our Hotel, col E+ = comps
    ws   = wb["Rates"]
    rows = list(ws.iter_rows(values_only=True))

    def _clean(v):
        if v is None: return None
        s = str(v).strip().lstrip("'")
        return s if s else None

    # Find header row (contains our hotel name or known comp names)
    our_col_idx  = None
    comp_col_idxs = []
    header_idx   = 4  # default row 5
    for i in range(2, min(10, len(rows))):
        row = rows[i]
        if not row: continue
        for j, cell in enumerate(row):
            v = _clean(cell)
            if not v: continue
            # Check if this cell is our hotel (first entry in comp_hotels)
            if comp_hotels and (v == comp_hotels[0] or comp_hotels[0].lower() in v.lower() or v.lower() in comp_hotels[0].lower()):
                our_col_idx = j
                header_idx  = i
                break
        if our_col_idx is not None:
            break

    if our_col_idx is None:
        # Fallback: assume col D (idx 3) is our hotel, col B (idx 1) date
        our_col_idx = 3
        header_idx  = 4

    # Find all comp columns from the same header row
    header_row = rows[header_idx]
    comp_col_idxs = []
    for j, cell in enumerate(header_row):
        if j == our_col_idx: continue
        v = _clean(cell)
        if not v: continue
        for full_name in comp_hotels[1:]:  # skip our hotel
            if (v == full_name or full_name.lower() in v.lower() or v.lower() in full_name.lower()):
                comp_col_idxs.append(j)
                break

    # Date column: look for col B (idx 1) containing dates
    date_col = 1

    records = []
    for row in rows[header_idx + 1:]:
        if not row or len(row) <= our_col_idx: continue
        dt_raw = row[date_col]
        dt = pd.to_datetime(dt_raw, errors="coerce") if dt_raw else pd.NaT
        if pd.isna(dt): continue

        # Market demand: col C (idx 2)
        dem_raw = row[2] if len(row) > 2 else None
        try:
            demand = round(float(str(dem_raw).rstrip("%")) * (1 if float(str(dem_raw).rstrip("%")) <= 1 else 0.01), 3) if dem_raw else None
        except (ValueError, TypeError):
            demand = None

        our_raw = row[our_col_idx] if our_col_idx < len(row) else None
        our_v   = _clean(our_raw)
        our_rate = None
        if our_v and "sold" not in our_v.lower() and not our_v.startswith("["):
            try: our_rate = float(our_v.split()[0].replace(",", "").replace("$", ""))
            except (ValueError, TypeError): pass

        # Compute median comp from comp columns
        comp_vals = []
        for ci in comp_col_idxs:
            if ci >= len(row): continue
            cv = _clean(row[ci])
            if not cv or "sold" in cv.lower() or cv.startswith("["): continue
            try: comp_vals.append(float(cv.split()[0].replace(",", "").replace("$", "")))
            except (ValueError, TypeError): pass
        med_comp = sorted(comp_vals)[len(comp_vals) // 2] if comp_vals else None

        records.append({
            "Day":           _clean(row[0]) if row[0] else None,
            "Date":          dt,
            "Our_Rate":      our_rate,
            "Median_Comp":   med_comp,
            "Rank":          None,
            "Market_Demand": demand,
            "Holiday":       None,
            "Event":         None,
        })

    df = pd.DataFrame(records).dropna(subset=["Date"]).sort_values("Date").reset_index(drop=True)
    df["Date"]        = pd.to_datetime(df["Date"]).dt.normalize()
    df["Our_Rate"]    = pd.to_numeric(df["Our_Rate"],    errors="coerce")
    df["Median_Comp"] = pd.to_numeric(df["Median_Comp"], errors="coerce")
    return df


def load_rates_comp(path, hotel_id: str = "hampton_lino_lakes") -> pd.DataFrame:
    """
    Reads Rates.xlsx Rates sheet — all comp hotel rates by date.
    Returns long-format DataFrame: Date, Hotel, Rate

    Supports Lighthouse format (all hotels on Rates sheet, row 5 = header,
    col B = Day, col C = Date, col D = Market demand, col E+ = hotels).
    Header row is always index 4 in Lighthouse exports.
    Hotel columns are mapped by name match against comp registry; if a name
    doesn't match, it is mapped positionally to the registry order as fallback.
    """
    comp_hotels, comp_short = get_comp_set(hotel_id)

    def _clean(v):
        """Strip leading apostrophe Excel text-prefix and whitespace."""
        if v is None: return None
        s = str(v).strip().lstrip("'")
        return s if s else None

    wb = openpyxl.load_workbook(_open(path), read_only=True, data_only=True)
    ws = wb["Rates"]
    rows = list(ws.iter_rows(values_only=True))

    # Header row: always index 4 in Lighthouse format (row 5).
    # Verify by checking that col 1 = "Day" and col 2 = "Date".
    # Fall back to scanning if layout differs.
    header_idx = 4
    for i in range(2, min(10, len(rows))):
        row = rows[i]
        if not row or len(row) < 3: continue
        v1 = _clean(row[1]); v2 = _clean(row[2])
        if v1 and v2 and "day" in v1.lower() and "date" in v2.lower():
            header_idx = i
            break

    headers = [_clean(h) for h in rows[header_idx]] if header_idx < len(rows) else []

    # Build hotel_cols: map column index -> short hotel name.
    # Fixed cols 0-3 are always: blank, Day, Date, Market demand — skip them.
    # Name match: case-insensitive substring. Positional fallback: any hotel
    # column that doesn't name-match gets the next unmatched registry entry in order.
    SKIP_COLS = {0, 1, 2, 3}
    name_matched    = {}
    unmatched_cols  = []
    already_matched = set()

    for i, h in enumerate(headers):
        if i in SKIP_COLS or not h:
            continue
        matched = False
        for full_name in comp_hotels:
            if full_name in already_matched:
                continue
            h_l = h.lower(); fn_l = full_name.lower()
            if h_l == fn_l or h_l in fn_l or fn_l in h_l:
                name_matched[i] = comp_short.get(full_name, full_name)
                already_matched.add(full_name)
                matched = True
                break
        if not matched:
            unmatched_cols.append(i)

    remaining_comps = [ch for ch in comp_hotels if ch not in already_matched]
    for col_idx, full_name in zip(unmatched_cols, remaining_comps):
        name_matched[col_idx] = comp_short.get(full_name, full_name)

    hotel_cols = name_matched

    # Detect date column: scan first data row for first column containing a parseable date
    # IDeaS layout: col C (idx 2) = date. Lighthouse layout: col B (idx 1) = date.
    date_col_idx  = 2   # IDeaS default
    demand_col_idx = 3  # IDeaS default
    for probe_row in rows[header_idx + 1:header_idx + 5]:
        if not probe_row: continue
        for j, cell in enumerate(probe_row):
            if cell is None: continue
            try:
                dt_probe = pd.to_datetime(cell, errors="coerce")
                if dt_probe is not pd.NaT and not pd.isna(dt_probe):
                    date_col_idx   = j
                    demand_col_idx = j + 1  # demand is always the column immediately after date
                    break
            except Exception:
                continue
        break

    records = []
    for row in rows[header_idx + 1:]:
        if not row or len(row) < 3:
            continue
        dt_raw = _clean(row[date_col_idx]) if date_col_idx < len(row) else None
        dt = pd.to_datetime(dt_raw, errors="coerce") if dt_raw else pd.NaT
        if pd.isna(dt):
            continue
        demand_raw = row[demand_col_idx] if demand_col_idx < len(row) else None
        try:
            dem_str = str(demand_raw).lstrip("'").rstrip("%")
            dem_val = float(dem_str)
            # Lighthouse stores demand as "59%" string or 0.59 float; IDeaS stores as 0.xx float
            demand = round(dem_val * 100, 1) if dem_val <= 1 else round(dem_val, 1)
        except (ValueError, TypeError):
            demand = None
        for col_idx, hotel_name in hotel_cols.items():
            rate_raw = row[col_idx] if col_idx < len(row) else None
            rate = _clean(rate_raw) if rate_raw is not None else None
            if rate and ("sold" in rate.lower() or rate.startswith("[")):
                rate = "Sold out"
            records.append({
                "Date":          dt,
                "Day":           _clean(row[1]) if len(row) > 1 else None,
                "Market_Demand": demand,
                "Hotel":         hotel_name,
                "Rate":          rate,
            })

    if not records:
        return pd.DataFrame(columns=["Date", "Day", "Market_Demand", "Hotel", "Rate", "Rate_Numeric"])

    df = pd.DataFrame(records).dropna(subset=["Date"])
    df = df[df["Rate"].notna()]
    df["Date"] = pd.to_datetime(df["Date"]).dt.normalize()
    df["Rate_Numeric"] = pd.to_numeric(df["Rate"], errors="coerce")
    return df.sort_values(["Date", "Hotel"]).reset_index(drop=True)


def load_rates_changes(path, vs_sheet: str = "vs. Yesterday", hotel_id: str = "hampton_lino_lakes") -> pd.DataFrame:
    """
    Reads one of the rate change comparison sheets.
    vs_sheet: "vs. Yesterday" | "vs. 3 days ago" | "vs. 7 days ago"
    Returns wide DataFrame: Date, Day, HotelShort__rate, HotelShort__change, ...
    """
    comp_hotels, comp_short = get_comp_set(hotel_id)

    def _clean(v):
        if v is None: return None
        s = str(v).strip().lstrip("'")
        return s if s else None

    wb = openpyxl.load_workbook(_open(path), read_only=True, data_only=True)
    ws = wb[vs_sheet]
    rows = list(ws.iter_rows(values_only=True))

    headers = [_clean(h) for h in rows[4]]

    # Build hotel_rate_cols: map col_idx -> (short_name, "rate"|"change")
    # vs. sheets layout: col 0=blank, 1=Day, 2=Date, 3=Market demand,
    # then hotel name at every other column starting at col 5 (col 4 is blank spacer).
    # Data rows: hotel rate at same col as name, change delta at col+1.
    # Name match first; positional fallback for any that don't match.
    SKIP_COLS = {0, 1, 2, 3}
    name_matched_h  = {}   # col_idx -> short_name (rate col only)
    unmatched_cols_h = []
    already_matched_h = set()

    for i, h in enumerate(headers):
        if i in SKIP_COLS or not h: continue
        matched = False
        for full_name in comp_hotels:
            if full_name in already_matched_h: continue
            h_l = h.lower(); fn_l = full_name.lower()
            if h_l == fn_l or h_l in fn_l or fn_l in h_l:
                name_matched_h[i] = comp_short.get(full_name, full_name)
                already_matched_h.add(full_name)
                matched = True
                break
        if not matched:
            unmatched_cols_h.append(i)

    remaining_h = [ch for ch in comp_hotels if ch not in already_matched_h]
    for col_idx, full_name in zip(unmatched_cols_h, remaining_h):
        name_matched_h[col_idx] = comp_short.get(full_name, full_name)

    # Build full hotel_rate_cols: rate at col_i, change at col_i+1
    hotel_rate_cols = {}
    for col_idx, short_name in name_matched_h.items():
        hotel_rate_cols[col_idx]     = (short_name, "rate")
        hotel_rate_cols[col_idx + 1] = (short_name, "change")

    records = []
    for row in rows[5:]:
        dt_raw = _clean(row[2]) if len(row) > 2 else None
        dt = pd.to_datetime(dt_raw, errors="coerce") if dt_raw else pd.NaT
        if pd.isna(dt):
            continue
        row_data = {"Date": dt, "Day": _clean(row[1]) if len(row) > 1 else None}
        for col_idx, (hotel, field) in hotel_rate_cols.items():
            val = row[col_idx] if col_idx < len(row) else None
            if field == "rate" and val is not None:
                v = _clean(val)
                val = "Sold out" if (v and ("sold" in v.lower() or v.startswith("["))) else v
            elif field == "change" and val is not None:
                try: val = float(str(val).lstrip("'"))
                except (ValueError, TypeError): val = None
            row_data[f"{hotel}__{field}"] = val
        records.append(row_data)

    df = pd.DataFrame(records).dropna(subset=["Date"])
    df["Date"] = pd.to_datetime(df["Date"]).dt.normalize()
    # Ensure every hotel has both __rate and __change columns (fill missing change with None)
    _, comp_short_map = get_comp_set(hotel_id)
    for short in comp_short_map.values():
        if f"{short}__rate" in df.columns and f"{short}__change" not in df.columns:
            df[f"{short}__change"] = None
    return df


# ═══════════════════════════════════════════════════════════════════════════════
# STR.XLSX  —  Competitive set STR report
# ═══════════════════════════════════════════════════════════════════════════════

def load_str_glance(path) -> dict:
    """
    Parses the STR Glance sheet (weekly + 28-day sections).
    Returns dict with 'period', 'weekly_df', '28day_df'
    """
    wb = openpyxl.load_workbook(_open(path), read_only=True, data_only=True)
    ws = wb["Glance"]
    rows = list(ws.iter_rows(values_only=True))

    period = rows[5][1] if rows[5][1] else "Unknown"

    # Column positions for Sun/Mon/Tue/Wed/Thu/Fri/Sat/Total
    # Each day has: [val, chg] pair — val_cols are the value columns, chg_cols are +1
    val_cols = [4, 7, 10, 13, 16, 19, 22, 25]
    chg_cols = [5, 8, 11, 14, 17, 20, 23, 26]
    days     = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Total"]

    def to_pct_chg_g(v):
        """Convert % change value — multiply by 100 if stored as decimal fraction."""
        if v is None: return None
        try:
            f = float(v)
            return round(f * 100, 1) if abs(f) < 5 else round(f, 1)
        except: return None

    def parse_section(occ_mine_r, occ_comp_r, mpi_r, adr_mine_r, adr_comp_r, ari_r,
                      rev_mine_r, rev_comp_r, rgi_r):
        def get(r, col): return rows[r][col] if len(rows[r]) > col else None
        data = []
        for i, (day, vc, cc) in enumerate(zip(days, val_cols, chg_cols)):
            occ_m = get(occ_mine_r, vc)
            occ_c = get(occ_comp_r, vc)
            data.append({
                "Day":              day,
                "Occ_Mine":         round(occ_m * 100, 1) if occ_m and occ_m < 2 else occ_m,
                "Occ_Mine_Chg":     to_pct_chg_g(get(occ_mine_r, cc)),
                "Occ_Comp":         round(occ_c * 100, 1) if occ_c and occ_c < 2 else occ_c,
                "Occ_Comp_Chg":     to_pct_chg_g(get(occ_comp_r, cc)),
                "MPI":              get(mpi_r, vc),
                "MPI_Chg":          to_pct_chg_g(get(mpi_r, cc)),
                "ADR_Mine":         get(adr_mine_r, vc),
                "ADR_Mine_Chg":     to_pct_chg_g(get(adr_mine_r, cc)),
                "ADR_Comp":         get(adr_comp_r, vc),
                "ADR_Comp_Chg":     to_pct_chg_g(get(adr_comp_r, cc)),
                "ARI":              get(ari_r, vc),
                "ARI_Chg":          to_pct_chg_g(get(ari_r, cc)),
                "RevPAR_Mine":      get(rev_mine_r, vc),
                "RevPAR_Mine_Chg":  to_pct_chg_g(get(rev_mine_r, cc)),
                "RevPAR_Comp":      get(rev_comp_r, vc),
                "RevPAR_Comp_Chg":  to_pct_chg_g(get(rev_comp_r, cc)),
                "RGI":              get(rgi_r, vc),
                "RGI_Chg":          to_pct_chg_g(get(rgi_r, cc)),
            })
        return pd.DataFrame(data)

    weekly_df = parse_section(9, 10, 11, 13, 14, 15, 17, 18, 19)
    day28_df  = parse_section(26, 27, 28, 30, 31, 32, 34, 35, 36)

    return {"period": period, "weekly": weekly_df, "28day": day28_df}


def load_str_segmentation(path) -> dict:
    """
    Parses the STR Segmentation Glance sheet.
    Layout:
      Row 6  : period label (col B)
      Row 8  : segment headers — Transient / Group / Contract / Total
      Row 10 : This Week section starts (9 rows: Occ mine/comp/idx, ADR mine/comp/idx, RevPAR mine/comp/idx)
      Row 22 : Running 28 Days section starts (same 9-row structure)
      Col D  : row labels
      Cols per segment (val, %chg):
        Transient  E(4) F(5)
        Group      G(6) H(7)
        Contract   I(8) J(9)
        Total      K(10) L(11)
    """
    wb = openpyxl.load_workbook(_open(path), read_only=True, data_only=True)

    # Try both common sheet names
    seg_sheet = None
    for name in ["Segmentation Glance", "Seg Glance", "Segmentation"]:
        if name in wb.sheetnames:
            seg_sheet = name
            break
    if seg_sheet is None:
        return None

    ws   = wb[seg_sheet]
    rows = list(ws.iter_rows(values_only=True))

    # Period label — row 6 (index 5), col B (index 1)
    period = rows[5][1] if len(rows) > 5 and rows[5][1] else "Unknown"

    # Exact column indices (0-based) confirmed from cell references:
    # D=3 (labels)
    # Transient: E=4 (val), F=5 (chg)
    # Group:     H=7 (val), I=8 (chg)
    # Contract:  L=11 (val), M=12 (chg)
    # Total:     P=15 (val), Q=16 (chg)
    # Each segment block: [label col][value col][% chg col]
    # Transient: D(lbl) E(val=4)  F(chg=5)
    # Group:     H(lbl) I(val=8)  J(chg=9)
    # Contract:  L(lbl) M(val=12) N(chg=13)
    # Total:     P(lbl) Q(val=16) R(chg=17)
    SEGS = [
        ("Transient",  4,  5),
        ("Group",      8,  9),
        ("Contract",  12, 13),
        ("Total",     16, 17),
    ]

    def safe(row_idx, col_idx):
        if row_idx >= len(rows): return None
        r = rows[row_idx]
        if col_idx >= len(r): return None
        v = r[col_idx]
        return v if v not in (None, "") else None

    def to_pct(v):
        if v is None: return None
        try:
            f = float(v)
            return round(f * 100, 1) if f < 2 else round(f, 1)
        except: return None

    def to_val(v):
        if v is None: return None
        try: return round(float(v), 2)
        except: return None

    def to_pct_chg(v):
        if v is None: return None
        try:
            f = float(v)
            return round(f * 100, 1) if abs(f) < 5 else round(f, 1)
        except: return None

    def parse_seg_section(base_row):
        """
        Parse one period block using absolute row indices (0-based).
        Confirmed from spreadsheet cell references:
          Occ  Mine/Comp/Index : base_row+0, base_row+1, base_row+2  (rows 10/11/12 → 9/10/11)
          ADR  Mine/Comp/Index : base_row+4, base_row+5, base_row+6  (rows 14/15/16 → 13/14/15)
          RevPAR Mine/Comp/Index: base_row+8, base_row+9, base_row+10 (rows 18/19/20 → 17/18/19)
        """
        records = []
        for seg_name, vc, cc in SEGS:
            occ_mine = to_pct(safe(base_row + 0, vc));  occ_mine_chg = to_pct_chg(safe(base_row + 0, cc))
            occ_comp = to_pct(safe(base_row + 1, vc));  occ_comp_chg = to_pct_chg(safe(base_row + 1, cc))
            mpi      = to_val(safe(base_row + 2, vc));  mpi_chg      = to_pct_chg(safe(base_row + 2, cc))
            adr_mine = to_val(safe(base_row + 4, vc));  adr_mine_chg = to_pct_chg(safe(base_row + 4, cc))
            adr_comp = to_val(safe(base_row + 5, vc));  adr_comp_chg = to_pct_chg(safe(base_row + 5, cc))
            ari      = to_val(safe(base_row + 6, vc));  ari_chg      = to_pct_chg(safe(base_row + 6, cc))
            rev_mine = to_val(safe(base_row + 8, vc));  rev_mine_chg = to_pct_chg(safe(base_row + 8, cc))
            rev_comp = to_val(safe(base_row + 9, vc));  rev_comp_chg = to_pct_chg(safe(base_row + 9, cc))
            rgi      = to_val(safe(base_row +10, vc));  rgi_chg      = to_pct_chg(safe(base_row +10, cc))
            records.append({
                "Segment":         seg_name,
                "Occ_Mine":        occ_mine,  "Occ_Mine_Chg":    occ_mine_chg,
                "Occ_Comp":        occ_comp,  "Occ_Comp_Chg":    occ_comp_chg,
                "MPI":             mpi,       "MPI_Chg":         mpi_chg,
                "ADR_Mine":        adr_mine,  "ADR_Mine_Chg":    adr_mine_chg,
                "ADR_Comp":        adr_comp,  "ADR_Comp_Chg":    adr_comp_chg,
                "ARI":             ari,       "ARI_Chg":         ari_chg,
                "RevPAR_Mine":     rev_mine,  "RevPAR_Mine_Chg": rev_mine_chg,
                "RevPAR_Comp":     rev_comp,  "RevPAR_Comp_Chg": rev_comp_chg,
                "RGI":             rgi,       "RGI_Chg":         rgi_chg,
            })
        return pd.DataFrame(records)

    # This Week starts row 10 (index 9); Running 28 starts row 27 (index 26)
    # Running 28 header is row 23 (index 22), data starts row 27 (index 26)
    weekly_df = parse_seg_section(9)
    day28_df  = parse_seg_section(26)

    return {"period": period, "weekly": weekly_df, "28day": day28_df}


def load_str_daily(path, metric: str = "Occ") -> pd.DataFrame:
    """
    Reads daily time-series from Occ, ADR, or RevPAR sheet.
    Returns: Day (int), My_Property, Comp_Set, Industry, Rank
    """
    wb = openpyxl.load_workbook(_open(path), read_only=True, data_only=True)
    ws = wb[metric]
    rows = list(ws.iter_rows(values_only=True))

    day_nums  = [v for v in rows[23][2:] if v is not None]
    mine_vals = [v for v in rows[24][2:] if v is not None]
    comp_vals = [v for v in rows[25][2:] if v is not None]
    ind_vals  = [v for v in rows[26][2:] if v is not None]

    n = min(len(day_nums), len(mine_vals), len(comp_vals), len(ind_vals))
    df = pd.DataFrame({
        "Day":         day_nums[:n],
        "My_Property": mine_vals[:n],
        "Comp_Set":    comp_vals[:n],
        "Industry":    ind_vals[:n],
    })

    try:
        rank_vals = [v for v in rows[50][2:] if v is not None]
        df["Rank"] = rank_vals[:n]
    except Exception:
        pass

    return df


def load_str_ranks(path) -> dict:
    """
    Reads the Current Week, Running 28, and MTD rank summaries from the
    Occ, ADR, and RevPAR sheets.
    The rank row is row 51 (index 50) — same as load_str_daily.
    The last 3 non-None numeric values in that row are Current Week, Run 28, MTD.
    """
    result = {}
    for metric, key in [("Occ", "occ"), ("ADR", "adr"), ("RevPAR", "revpar")]:
        try:
            wb   = openpyxl.load_workbook(_open(path), read_only=True, data_only=True)
            ws   = wb[metric]
            rows = list(ws.iter_rows(values_only=True))

            # Try rows 50 and 51 (indices 49 and 50) — pick the one with numeric values
            rank_row = None
            for idx in [50, 49, 51]:
                if idx >= len(rows):
                    continue
                candidates = [v for v in rows[idx] if v is not None and str(v).strip() not in ("", "Your rank", "Rank")]
                # Must contain numeric-looking values like "7 of 8" or integers
                if len(candidates) >= 3:
                    rank_row = candidates
                    break

            if rank_row and len(rank_row) >= 3:
                def fmt_rank(v):
                    s = str(v).strip()
                    # Already formatted as "X of Y"
                    if "of" in s.lower():
                        return s
                    # Raw integer — format as "X of ?"
                    try:
                        return str(int(float(s)))
                    except Exception:
                        return s

                result[key] = {
                    "week":  fmt_rank(rank_row[-3]),
                    "run28": fmt_rank(rank_row[-2]),
                    "mtd":   fmt_rank(rank_row[-1]),
                }
            else:
                result[key] = {"week": "—", "run28": "—", "mtd": "—"}
        except Exception:
            result[key] = {"week": "—", "run28": "—", "mtd": "—"}
    return result




# ═══════════════════════════════════════════════════════════════════════════════
# SRP_PACE.XLSX  --  Booking Situation Report (segment-level pace vs STLY)
# 13 segments x 11 sub-columns each, daily rows for ~2 years
# ===============================================================================

def load_srp_pace(path) -> pd.DataFrame:
    """
    Reads SRP Pace.xlsx.  Handles three formats:

    Format C (Hilton wide — detected by Row 6 having MCAT segment names):
      Row 6 = segment names (BAR, CMP, CMTG, …) starting at col 4, every 11 cols
      Row 7 = measure sub-headers
      Row 8+ = one row per stay date; within each segment block:
        +2=OTB, +3=STLY_OTB, +5=Revenue, +6=STLY_Revenue, +8=ADR, +9=STLY_ADR

    Format B (IHG pivoted): col3 == "Measure Names"
    Format A (IHG wide):    original 29-col layout
    """
    import openpyxl, warnings
    warnings.filterwarnings("ignore")

    wb = openpyxl.load_workbook(_open(path), data_only=True)
    ws = wb.active

    def _sf(v):
        if v is None: return 0.0
        try: return float(str(v).replace(",", ""))
        except: return 0.0

    # ── Detect Format C: Hilton wide-format ──────────────────────────────────
    HILTON_MCATS = {"BAR","CMP","CMTG","CNR","CONS","CONV","DISC","GOV","GT",
                    "HOU","IT","LNR","MKT","PERM","SMRF","SMERF"}
    row6 = [ws.cell(6, c).value for c in range(1, ws.max_column + 1)]
    seg_cols = {}  # segment name → 1-based start col
    for i, v in enumerate(row6):
        if v and str(v).strip() in HILTON_MCATS:
            seg_cols[str(v).strip()] = i + 1

    if seg_cols:
        # Format C — wide format, one row per stay date
        # Within each segment block (start col s):
        #   s+2=OTB, s+3=STLY_OTB, s+5=Revenue, s+6=STLY_Revenue, s+8=ADR, s+9=STLY_ADR
        records = []
        for r in range(8, ws.max_row + 1):
            date_val = ws.cell(r, 1).value
            if not date_val or not isinstance(date_val, (pd.Timestamp, __import__('datetime').datetime, __import__('datetime').date)):
                try:
                    date_val = pd.to_datetime(date_val)
                    if pd.isna(date_val):
                        continue
                except Exception:
                    continue
            date_ts = pd.Timestamp(date_val).normalize()
            for seg, s in seg_cols.items():
                otb      = _sf(ws.cell(r, s + 2).value)
                stly_otb = _sf(ws.cell(r, s + 3).value)
                rev      = _sf(ws.cell(r, s + 5).value)
                stly_rev = _sf(ws.cell(r, s + 6).value)
                adr      = _sf(ws.cell(r, s + 8).value)  if ws.cell(r, s + 8).value else (rev / otb if otb > 0 else 0.0)
                stly_adr = _sf(ws.cell(r, s + 9).value)  if ws.cell(r, s + 9).value else (stly_rev / stly_otb if stly_otb > 0 else 0.0)
                # Normalise SMERF→SMRF for consistency with SEGS_ORDER
                seg_key = "SMRF" if seg == "SMERF" else seg
                records.append({
                    "Date":         date_ts,
                    "Segment":      seg_key,
                    "OTB":          otb,
                    "STLY_OTB":     stly_otb,
                    "Var_OTB":      otb - stly_otb,
                    "Revenue":      rev,
                    "STLY_Revenue": stly_rev,
                    "Var_Revenue":  rev - stly_rev,
                    "ADR":          adr,
                    "STLY_ADR":     stly_adr,
                    "Var_ADR":      adr - stly_adr,
                })
        df = pd.DataFrame(records)
        if df.empty:
            return df
        df["Date"] = pd.to_datetime(df["Date"])
        return df.sort_values(["Date", "Segment"]).reset_index(drop=True)

    # ── Format B / A: IHG segment-mapped formats ─────────────────────────────
    def _ym_to_date(ym):
        try:
            parts = str(ym).split("-M")
            return pd.Timestamp(year=int(parts[0]), month=int(parts[1]), day=1)
        except: return pd.NaT

    h3 = str(ws.cell(1, 3).value or "").strip()

    if h3 == "Measure Names":
        # Format B — pivoted IHG
        MEASURE_OTB     = "OTB Rooms By Segment"
        MEASURE_STLY    = "Last Year OTB Rooms By Segment"
        MEASURE_REV     = "OTB Revenue By Segment - USD"
        MEASURE_STLYREV = "Last Year OTB Revenue By Segment - USD"
        MEASURE_PROJR   = "Projected Rooms By Segment"
        MEASURE_PROJREV = "Projected Revenue By Segment - USD"

        from collections import defaultdict
        buckets = defaultdict(lambda: defaultdict(float))

        for r in range(2, ws.max_row + 1):
            measure  = ws.cell(r, 3).value
            sales_sg = ws.cell(r, 4).value
            ym       = ws.cell(r, 5).value
            val      = ws.cell(r, 7).value
            if not measure or not ym or not sales_sg: continue
            measure_s = str(measure).strip()
            sales_s   = str(sales_sg).strip()
            mapped    = _IHG_SEGMENT_MAP.get(sales_s)
            if not mapped: continue
            key = (str(ym).strip(), mapped)
            v   = _sf(val)
            if measure_s == MEASURE_OTB:      buckets[key]["otb"]       += v
            elif measure_s == MEASURE_STLY:   buckets[key]["stly_otb"]  += v
            elif measure_s == MEASURE_REV:    buckets[key]["rev"]        += v
            elif measure_s == MEASURE_STLYREV:buckets[key]["stly_rev"]   += v
            elif measure_s == MEASURE_PROJR:  buckets[key]["proj_rooms"] += v
            elif measure_s == MEASURE_PROJREV:buckets[key]["proj_rev"]   += v

        records = []
        for (ym, seg), b in sorted(buckets.items()):
            date_val = _ym_to_date(ym)
            if pd.isna(date_val): continue
            otb=b["otb"]; stly=b["stly_otb"]; rev=b["rev"]; srev=b["stly_rev"]
            proj_r=b.get("proj_rooms", otb); proj_v=b.get("proj_rev", rev)
            adr  = rev  / otb   if otb  > 0 else 0.0
            sadr = srev / stly  if stly > 0 else 0.0
            padr = proj_v / proj_r if proj_r > 0 else 0.0
            records.append({"Date": date_val, "Segment": seg,
                "OTB": otb, "STLY_OTB": stly, "Var_OTB": otb - stly,
                "Revenue": rev, "STLY_Revenue": srev, "Var_Revenue": rev - srev,
                "ADR": adr, "STLY_ADR": sadr, "Var_ADR": adr - sadr,
                "Projected_Rooms": proj_r, "Projected_Rev": proj_v, "Projected_ADR": padr})
    else:
        # Format A — original IHG wide format
        from collections import defaultdict
        buckets = defaultdict(lambda: {"otb": 0., "stly_otb": 0., "rev": 0., "stly_rev": 0., "proj_rooms": 0., "proj_rev": 0.})
        for r in range(2, ws.max_row + 1):
            ym       = ws.cell(r, 2).value
            sales_sg = ws.cell(r, 5).value
            if not ym or not sales_sg: continue
            mapped = _IHG_SEGMENT_MAP.get(str(sales_sg).strip())
            if not mapped: continue
            key = (str(ym).strip(), mapped)
            b   = buckets[key]
            b["otb"]       += _sf(ws.cell(r,  6).value)
            b["stly_otb"]  += _sf(ws.cell(r,  7).value)
            b["rev"]       += _sf(ws.cell(r,  9).value)
            b["stly_rev"]  += _sf(ws.cell(r, 10).value)
            b["proj_rooms"]+= _sf(ws.cell(r, 15).value)
            b["proj_rev"]  += _sf(ws.cell(r, 18).value)
        records = []
        for (ym, seg), b in sorted(buckets.items()):
            date_val = _ym_to_date(ym)
            if pd.isna(date_val): continue
            otb=b["otb"]; stly=b["stly_otb"]; rev=b["rev"]; srev=b["stly_rev"]
            proj_r=b["proj_rooms"]; proj_v=b["proj_rev"]
            adr  = rev  / otb   if otb  > 0 else 0.0
            sadr = srev / stly  if stly > 0 else 0.0
            padr = proj_v / proj_r if proj_r > 0 else 0.0
            records.append({"Date": date_val, "Segment": seg,
                "OTB": otb, "STLY_OTB": stly, "Var_OTB": otb - stly,
                "Revenue": rev, "STLY_Revenue": srev, "Var_Revenue": rev - srev,
                "ADR": adr, "STLY_ADR": sadr, "Var_ADR": adr - sadr,
                "Projected_Rooms": proj_r, "Projected_Rev": proj_v, "Projected_ADR": padr})

    df = pd.DataFrame(records)
    if df.empty:
        return df
    df["Date"] = pd.to_datetime(df["Date"])
    return df.sort_values(["Date", "Segment"]).reset_index(drop=True)

# ── data_loader version — increment when loader logic changes ────────────────
_DL_VERSION = "5.9-ihg-forecast-crash-fix"

# ═══════════════════════════════════════════════════════════════════════════════
# IHG-SPECIFIC LOADERS  —  Holiday Inn Express Superior and future IHG hotels
# ═══════════════════════════════════════════════════════════════════════════════

# Hotels that use IHG report formats instead of Hilton formats
IHG_HOTELS = {"holiday_inn_express_superior", "hotel_indigo_rochester"}

# IHG segment rollup: Sales Segment → RevPar MD display segment
_IHG_SEGMENT_MAP = {
    "GSO Managed Negotiated":           "CORPORATE",
    "Hotel Locally Managed Negotiated": "CORPORATE",
    "GSO Managed Government":           "GOVERNMENT",
    "Hotel Locally Managed Government": "GOVERNMENT",
    "Group":                            "GROUP",
    "Group Corporate":                  "GROUP",
    "Group Government":                 "GROUP",
    "Group Leisure":                    "GROUP",
    "Fenced":                           "RETAIL",
    "Unfenced":                         "RETAIL",
    "Opaque":                           "RETAIL",
    "Extended Stay":                    "RETAIL",
    "Tactical Marketing":               "RETAIL",
    "Membership Marketing":             "RETAIL",
    "GSO Managed Membership Marketing": "RETAIL",
    "Other":                            "RETAIL",
    "GSO Managed Wholesale":            "WHOLESALE",
    "Hotel Locally Managed Wholesale":  "WHOLESALE",
}


def _parse_ihg_date(val) -> pd.Timestamp:
    """Parse IHG StrategicAnalysis date strings like '16Mar2026' or '01Mar2027'."""
    if val is None:
        return pd.NaT
    s = str(val).strip()
    try:
        return pd.to_datetime(s, format="%d%b%Y")
    except Exception:
        try:
            return pd.to_datetime(s)
        except Exception:
            return pd.NaT


def load_ihg_year(path) -> pd.DataFrame:
    """
    Reads Data_Glance.xlsx (IHG).
    Row 1 = headers. Data starts row 2.
    Returns DataFrame with same schema as load_year() so all existing
    tabs consume it transparently.

    Column mapping:
      Col 1  Date                    → Date (M/D/YYYY string)
      Col 3  Occ %                   → Forecast_Current (occ % — also used for past actuals)
      Col 4  ADR                     → ADR
      Col 5  BFR                     → BFR
      Col 6  AC                      → capacity (used for null-check only)
      Col 8  Total Rooms Committed   → OTB
      Col 9  Actual rooms sold LY    → OTB_LY
      Col 15 Rooms sold              → Forecast_Rooms (past dates only; SA merges future)

    Revenue_OTB and Revenue_LY are NOT in this file — populated later via
    merge with StrategicAnalysis and budget file respectively.
    """
    import openpyxl, warnings
    warnings.filterwarnings("ignore")

    wb = openpyxl.load_workbook(_open(path), data_only=True)
    ws = wb.active

    def _safe_float(v):
        try:
            return float(v) if v is not None else None
        except (ValueError, TypeError):
            return None

    records = []
    for r in range(2, ws.max_row + 1):
        raw_date = ws.cell(r, 1).value
        if not raw_date:
            continue
        try:
            date_val = pd.to_datetime(str(raw_date))
        except Exception:
            continue

        records.append({
            "Date":             date_val,
            "OTB":              _safe_float(ws.cell(r, 8).value),
            "OTB_LY":           _safe_float(ws.cell(r, 9).value),
            "Forecast_Rooms":   _safe_float(ws.cell(r, 15).value),
            "Forecast_Current": _safe_float(ws.cell(r, 3).value),
            "ADR":              _safe_float(ws.cell(r, 4).value),
            "BFR":              _safe_float(ws.cell(r, 5).value),
            "Revenue_OTB":      None,   # filled by merge with SA
            "Revenue_LY":       None,   # filled by merge with budget
        })

    df = pd.DataFrame(records)
    if df.empty:
        return df
    df["Date"] = pd.to_datetime(df["Date"])
    return df.sort_values("Date").reset_index(drop=True)


def load_ihg_strategic(path, total_rooms: int = 84) -> dict:
    """
    Reads StrategicAnalysis.xlsx (IHG).
    Row 1 = section headers (merged). Row 2 = column headers. Data starts row 3.
    Date format: '16Mar2026' (no separators).

    Returns dict with three keys:

    'year_supplement': DataFrame for merging into year df
        Date, Forecast_Rooms, Forecast_Current, Revenue_OTB, Our_Rate

    'rates': DataFrame for Rate Surveillance tab
        Date, Our_Rate (col 52),
        + one column per comp hotel named exactly as in _COMP_REGISTRY
          (cols 72–78 from the file)

    'pickup': DataFrame matching pickup schema
        Date, Pickup1 (col 13 Pickup Since Yesterday),
        Pickup7 (col 14 Pick up as of [date])

    Column reference (row 2 headers, 1-indexed):
      Col  2  Date
      Col  9  RC (remaining capacity — NOT used for OTB)
      Col 13  Pickup Since Yesterday          → Pickup1
      Col 14  Pick up as of [date]            → Pickup7
      Col 15  Forecast Occ%                   → Forecast_Current
      Col 47  ADR
      Col 50  Revenue (1000s)                 → Revenue_OTB (×1000)
      Col 51  RevPAR
      Col 52  Deployed BFR                    → Our_Rate
      Col 72  Americinn By Wyndham Duluth
      Col 73  Best Western Bridgeview Motor Superior
      Col 74  Comfort Inn West Duluth
      Col 75  Hampton Inn Duluth
      Col 76  Hampton Inn Superior Duluth Wi
      Col 77  Holiday Inn Express & Suites Duluth North - Miller Hill
      Col 78  La Quinta Inn Suites Duluth
    """
    import openpyxl, warnings
    warnings.filterwarnings("ignore")

    wb = openpyxl.load_workbook(_open(path), data_only=True)
    ws = wb.active

    def _safe_float(v):
        try:
            return float(v) if v is not None and str(v).strip() not in ("-", "", "N", "Y") else None
        except (ValueError, TypeError):
            return None

    # Read comp hotel headers from row 2 cols 72-78
    comp_cols = {}  # col_index → hotel name as it appears in file
    for c in range(72, 79):
        hname = ws.cell(2, c).value
        if hname:
            comp_cols[c] = str(hname).strip()

    year_supp_records = []
    rates_records     = []
    pickup_records    = []

    for r in range(3, ws.max_row + 1):
        raw_date = ws.cell(r, 2).value
        if not raw_date:
            continue
        date_val = _parse_ihg_date(raw_date)
        if pd.isna(date_val):
            continue

        fc_occ  = _safe_float(ws.cell(r, 15).value)
        rev_raw = _safe_float(ws.cell(r, 50).value)

        # ── Year supplement ──
        fc_rooms = round(fc_occ / 100.0 * total_rooms) if fc_occ is not None else None
        year_supp_records.append({
            "Date":             date_val,
            "Forecast_Rooms":   fc_rooms,
            "Forecast_Current": fc_occ,
            "Revenue_OTB":      rev_raw * 1000.0 if rev_raw is not None else None,
            "Our_Rate":         _safe_float(ws.cell(r, 52).value),
        })

        # ── Rates ──
        rate_row = {"Date": date_val, "Our_Rate": _safe_float(ws.cell(r, 52).value)}
        for c, hname in comp_cols.items():
            rate_row[hname] = _safe_float(ws.cell(r, c).value)
        rates_records.append(rate_row)

        # ── Pickup ──
        pickup_records.append({
            "Date":    date_val,
            "Pickup1": _safe_float(ws.cell(r, 13).value),
            "Pickup7": _safe_float(ws.cell(r, 14).value),
        })

    def _make_df(records):
        df = pd.DataFrame(records)
        if not df.empty:
            df["Date"] = pd.to_datetime(df["Date"])
            df = df.sort_values("Date").reset_index(drop=True)
        return df

    return {
        "year_supplement": _make_df(year_supp_records),
        "rates":           _make_df(rates_records),
        "pickup":          _make_df(pickup_records),
    }


def load_ihg_segment(path) -> pd.DataFrame:
    """
    Reads Property_Segment_Data.xlsx (IHG).
    Row 1 = headers. Data starts row 2.
    Monthly data keyed by Year_Month (e.g. '2026-M03').

    IHG Sales Segments are rolled up to 5 RevPar MD display segments
    via _IHG_SEGMENT_MAP. Rows with unmapped segments are silently skipped.

    Returns DataFrame with same schema as load_srp_pace() plus IHG-only
    Projected columns, so the adapted SRP Pace tab can use them:
      Date            — first day of the month (datetime)
      Segment         — CORPORATE / GOVERNMENT / GROUP / RETAIL / WHOLESALE
      OTB             — sum of OTB Rooms By Segment
      STLY_OTB        — sum of Last Year OTB Rooms By Segment
      Var_OTB         — OTB - STLY_OTB
      Revenue         — sum of OTB Revenue By Segment - USD
      STLY_Revenue    — sum of Last Year OTB Revenue By Segment - USD
      Var_Revenue     — Revenue - STLY_Revenue
      ADR             — Revenue / OTB (weighted; 0 if OTB=0)
      STLY_ADR        — STLY_Revenue / STLY_OTB (weighted; 0 if STLY_OTB=0)
      Var_ADR         — ADR - STLY_ADR
      Projected_Rooms — sum of Projected Rooms By Segment
      Projected_Rev   — sum of Projected Revenue By Segment - USD
      Projected_ADR   — Projected_Rev / Projected_Rooms (weighted)

    Column reference (row 1 headers, 1-indexed):
      Col  2  Year_Month
      Col  3  Market Segment
      Col  5  Sales Segment
      Col  6  OTB Rooms By Segment
      Col  7  Last Year OTB Rooms By Segment
      Col  9  OTB Revenue By Segment - USD
      Col 10  Last Year OTB Revenue By Segment - USD
      Col 15  Projected Rooms By Segment
      Col 18  Projected Revenue By Segment - USD
    """
    import openpyxl, warnings
    warnings.filterwarnings("ignore")

    wb = openpyxl.load_workbook(_open(path), data_only=True)
    ws = wb.active

    def _safe_float(v):
        try:
            f = float(v) if v is not None else 0.0
            return f if f == f else 0.0   # NaN guard
        except (ValueError, TypeError):
            return 0.0

    def _ym_to_date(ym_str: str) -> pd.Timestamp:
        """Convert '2026-M03' to 2026-03-01."""
        try:
            parts = str(ym_str).split("-M")
            return pd.Timestamp(year=int(parts[0]), month=int(parts[1]), day=1)
        except Exception:
            return pd.NaT

    # Accumulate by (Year_Month, mapped_segment)
    from collections import defaultdict
    buckets = defaultdict(lambda: {
        "otb": 0.0, "stly_otb": 0.0,
        "rev": 0.0, "stly_rev": 0.0,
        "proj_rooms": 0.0, "proj_rev": 0.0,
    })

    # Detect format from header row:
    #   Format A (old 29-col): col2=Year_Month, col5=Sales Segment, col6=OTB, col7=STLY_OTB,
    #                          col9=OTB Rev, col10=STLY Rev, col15=Proj Rooms, col18=Proj Rev
    #   Format B (new 7-col pivoted): col3=Measure Name, col4=Sales Segment, col5=Year_Month,
    #                                 col7=Measure Value (comma-formatted string)
    _h1 = ws.cell(1, 1).value or ""
    _h3 = ws.cell(1, 3).value or ""
    _fmt_b = "Measure" in str(_h3) or "Measure" in str(_h1)

    # Format B measure → bucket field mapping
    _MEASURE_TO_FIELD = {
        "OTB Rooms By Segment":                   "otb",
        "Last Year OTB Rooms By Segment":          "stly_otb",
        "OTB Revenue By Segment - USD":            "rev",
        "Last Year OTB Revenue By Segment - USD":  "stly_rev",
        "Projected Rooms By Segment":              "proj_rooms",
        "Projected Revenue By Segment - USD":      "proj_rev",
    }

    def _sf_str(v):
        """Parse comma-formatted string values like '1,843.99' used in Format B."""
        try:
            return float(str(v).replace(",", "")) if v is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    for r in range(2, ws.max_row + 1):
        if _fmt_b:
            # Format B: col3=Measure, col4=Sales Segment, col5=Year_Month, col7=Value
            measure  = ws.cell(r, 3).value
            sales_sg = ws.cell(r, 4).value
            ym       = ws.cell(r, 5).value
            val      = ws.cell(r, 7).value
            if not measure or not sales_sg or not ym:
                continue
            field  = _MEASURE_TO_FIELD.get(str(measure).strip())
            mapped = _IHG_SEGMENT_MAP.get(str(sales_sg).strip())
            if not field or not mapped:
                continue
            buckets[(str(ym).strip(), mapped)][field] += _sf_str(val)
        else:
            # Format A: col2=Year_Month, col5=Sales Segment
            ym       = ws.cell(r, 2).value
            sales_sg = ws.cell(r, 5).value
            if not ym or not sales_sg:
                continue
            mapped = _IHG_SEGMENT_MAP.get(str(sales_sg).strip())
            if not mapped:
                continue
            key = (str(ym).strip(), mapped)
            b   = buckets[key]
            b["otb"]        += _safe_float(ws.cell(r,  6).value)
            b["stly_otb"]   += _safe_float(ws.cell(r,  7).value)
            b["rev"]        += _safe_float(ws.cell(r,  9).value)
            b["stly_rev"]   += _safe_float(ws.cell(r, 10).value)
            b["proj_rooms"] += _safe_float(ws.cell(r, 15).value)
            b["proj_rev"]   += _safe_float(ws.cell(r, 18).value)

    records = []
    for (ym, seg), b in sorted(buckets.items()):
        date_val = _ym_to_date(ym)
        if pd.isna(date_val):
            continue

        otb      = b["otb"]
        stly_otb = b["stly_otb"]
        rev      = b["rev"]
        stly_rev = b["stly_rev"]
        proj_r   = b["proj_rooms"]
        proj_rev = b["proj_rev"]

        adr       = rev / otb            if otb      > 0 else 0.0
        stly_adr  = stly_rev / stly_otb  if stly_otb > 0 else 0.0
        proj_adr  = proj_rev / proj_r    if proj_r   > 0 else 0.0

        records.append({
            "Date":           date_val,
            "Segment":        seg,
            "OTB":            otb,
            "STLY_OTB":       stly_otb,
            "Var_OTB":        otb - stly_otb,
            "Revenue":        rev,
            "STLY_Revenue":   stly_rev,
            "Var_Revenue":    rev - stly_rev,
            "ADR":            adr,
            "STLY_ADR":       stly_adr,
            "Var_ADR":        adr - stly_adr,
            "Projected_Rooms":proj_r,
            "Projected_Rev":  proj_rev,
            "Projected_ADR":  proj_adr,
        })

    df = pd.DataFrame(records)
    if df.empty:
        return df
    df["Date"] = pd.to_datetime(df["Date"])
    return df.sort_values(["Date", "Segment"]).reset_index(drop=True)




def load_ihg_budget(path) -> pd.DataFrame:
    """
    Reads the IHG Budget.xlsx file.

    Layout (3 header rows, data from row 4):
      Row 1: Title
      Row 2: Section headers ("Budget", "Actuals / Definites Last Year")
      Row 3: Column headers
        Col A: Date  ("Thursday, January 1, 2026")
        Col B: Room  (total rooms = capacity)
        Col C: Occ Rooms  (Budget)
        Col D: Occ%       (Budget)
        Col E: ADR        (Budget)
        Col F: RevPAR     (Budget)
        Col G: Revenue    (Budget)
        Col H: Occ Rooms  (LY Actuals)
        Col I: Occ%       (LY)
        Col J: ADR        (LY)
        Col K: RevPAR     (LY)
        Col L: Revenue    (LY)  ← Revenue_LY

    Returns same schema as load_budget() PLUS Revenue_LY column.
    """
    wb = openpyxl.load_workbook(_open(path), data_only=True)
    preferred = ["Sheet1", "Budget"]
    sheet_name = next((n for n in preferred if n in wb.sheetnames), wb.sheetnames[0])

    # Read all 12 columns with header=2 (row 3 as header, data from row 4)
    df = pd.read_excel(_open(path), sheet_name=sheet_name, header=2, engine="openpyxl")

    # Need at least 12 columns (A-L); pad if fewer
    while len(df.columns) < 12:
        df[f"_pad_{len(df.columns)}"] = None

    # Extract budget columns (A-G) and LY columns (H-L)
    budget_cols = df.iloc[:, :7].copy()
    budget_cols.columns = ["Date_Raw", "Rooms", "Occ_Rooms", "Occ_Pct", "ADR", "RevPAR", "Revenue"]

    ly_cols = df.iloc[:, 7:12].copy()
    ly_cols.columns = ["LY_Occ_Rooms", "LY_Occ_Pct", "LY_ADR", "LY_RevPAR", "Revenue_LY"]

    # Combine
    combined = pd.concat([budget_cols, ly_cols], axis=1)

    # Parse dates — IHG uses "Thursday, January 1, 2026" format
    combined["Date"] = pd.to_datetime(combined["Date_Raw"], errors="coerce")
    combined = combined.dropna(subset=["Date"])
    combined = combined[combined["Date_Raw"] != "Total"].reset_index(drop=True)
    combined = combined.sort_values("Date").reset_index(drop=True)

    # Coerce numerics
    for col in ["Rooms", "Occ_Rooms", "Occ_Pct", "ADR", "RevPAR", "Revenue",
                "LY_Occ_Rooms", "LY_Occ_Pct", "LY_ADR", "LY_RevPAR", "Revenue_LY"]:
        combined[col] = pd.to_numeric(combined[col], errors="coerce")

    return combined[["Date", "Rooms", "Occ_Rooms", "Occ_Pct", "ADR", "RevPAR", "Revenue",
                      "LY_Occ_Rooms", "LY_Occ_Pct", "LY_ADR", "LY_RevPAR", "Revenue_LY"]]


def load_ihg_dss_property(path) -> pd.DataFrame:
    """
    Reads the IHG DSS Property Detail Report — pivoted format.

    Layout (Col A = Measure Name, Col B = Year_Month, Col C = sno, Col D = Value):
      Each row is one metric for one month.
      Year_Month format: "2026-M03"

    Auto-detects the correct sheet by scanning for "Projected Rooms By Segment"
    in column A. Falls back to first sheet if not found.

    Key measures extracted:
      "Projected Rooms By Segment"              → Projected_Rooms
      "Projected Revenue By Segment - USD"      → Projected_Rev
      "Last Year Sold Rooms By Segment"         → STLY_OTB
      "Last Year Total Revenue By Segment - USD"→ STLY_Revenue

    Returns DataFrame with same schema as load_ihg_segment() output so
    _build_ihg_daily_forecast() can consume it unchanged.
    """
    import openpyxl, warnings, re
    warnings.filterwarnings("ignore")

    wb = openpyxl.load_workbook(_open(path), data_only=True)

    # Find the "Property Data" sheet (contains pivoted hotel totals)
    # Sheet name priority: "Property Data" → scan all sheets for "Projected Rooms By Segment" in col A
    target_ws = wb["Property Data"] if "Property Data" in wb.sheetnames else None
    if target_ws is None:
        for sname in wb.sheetnames:
            ws = wb[sname]
            for r in range(1, min(ws.max_row + 1, 50)):
                v = ws.cell(r, 1).value
                if v and "Projected Rooms By Segment" in str(v):
                    target_ws = ws
                    break
            if target_ws is not None:
                break
    if target_ws is None:
        return pd.DataFrame()

    def _sf(v):
        """Parse value — handles both numeric and comma-formatted strings like '1,843.99'."""
        if v is None:
            return None
        try:
            return float(str(v).replace(",", ""))
        except (ValueError, TypeError):
            return None

    def _ym_to_date(ym_str: str) -> pd.Timestamp:
        try:
            parts = str(ym_str).split("-M")
            return pd.Timestamp(year=int(parts[0]), month=int(parts[1]), day=1)
        except Exception:
            return pd.NaT

    # Map measure name → field key
    MEASURE_MAP = {
        "Projected Rooms By Segment":               "proj_rooms",
        "Projected Revenue By Segment - USD":       "proj_rev",
        "Last Year Sold Rooms By Segment":          "stly_rooms",
        "Last Year Total Revenue By Segment - USD": "stly_rev",
    }

    # Collect {year_month: {field: value}}
    from collections import defaultdict
    monthly = defaultdict(dict)

    for r in range(2, target_ws.max_row + 1):
        measure = target_ws.cell(r, 1).value
        ym      = target_ws.cell(r, 2).value
        val     = target_ws.cell(r, 4).value
        if not measure or not ym:
            continue
        measure_str = str(measure).strip()
        field = MEASURE_MAP.get(measure_str)
        if field and val is not None:
            v = _sf(val)
            if v is not None:
                monthly[str(ym).strip()][field] = v

    if not monthly:
        return pd.DataFrame()

    records = []
    for ym, fields in sorted(monthly.items()):
        date_val = _ym_to_date(ym)
        if pd.isna(date_val):
            continue

        proj_rooms = fields.get("proj_rooms", 0.0)
        proj_rev   = fields.get("proj_rev",   0.0)
        stly_rooms = fields.get("stly_rooms", 0.0)
        stly_rev   = fields.get("stly_rev",   0.0)
        proj_adr   = proj_rev / proj_rooms if proj_rooms > 0 else 0.0
        stly_adr   = stly_rev / stly_rooms if stly_rooms > 0 else 0.0

        records.append({
            "Date":           date_val,
            "Segment":        "TOTAL",
            "OTB":            proj_rooms,   # use proj as OTB proxy for months with no OTB data
            "STLY_OTB":       stly_rooms,
            "Var_OTB":        proj_rooms - stly_rooms,
            "Revenue":        proj_rev,
            "STLY_Revenue":   stly_rev,
            "Var_Revenue":    proj_rev - stly_rev,
            "ADR":            proj_adr,
            "STLY_ADR":       stly_adr,
            "Var_ADR":        proj_adr - stly_adr,
            "Projected_Rooms":proj_rooms,
            "Projected_Rev":  proj_rev,
            "Projected_ADR":  proj_adr,
        })

    df = pd.DataFrame(records)
    if df.empty:
        return df
    df["Date"] = pd.to_datetime(df["Date"])
    return df.sort_values("Date").reset_index(drop=True)


def load_lighthouse_events(path) -> pd.DataFrame:
    """
    Load Lighthouse Events xlsx for IHG hotels.
    Robustly detects header row and column positions by name rather than
    assuming fixed column order or a specific sheet name.
    Returns a DataFrame with columns:
        Date, Event, Category, Location, Duration_Days, Start_Date, End_Date
    """
    try:
        wb = openpyxl.load_workbook(_open(path), data_only=True)
        # Prefer any sheet whose name contains "event" (case-insensitive), fallback to first sheet
        sheet_name = next(
            (s for s in wb.sheetnames if "event" in s.lower()),
            wb.sheetnames[0]
        )
        ws = wb[sheet_name]

        # ── Detect header row (scan first 15 rows for date/name/event keywords) ──
        col_map = {}
        COL_ALIASES = {
            "start_date": ["start date", "start_date", "startdate", "start"],
            "end_date":   ["end date",   "end_date",   "enddate",   "end"],
            "name":       ["name", "event name", "event_name", "title", "event"],
            "location":   ["location", "venue", "city"],
            "category":   ["category", "type", "event type", "event_type"],
        }
        header_row_idx = 0  # default fallback
        all_rows = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
        for row_i, row in enumerate(all_rows):
            matched = {}
            for col_i, cell in enumerate(row):
                if cell is None:
                    continue
                cell_str = str(cell).strip().lower()
                for canonical, aliases in COL_ALIASES.items():
                    if cell_str in aliases and canonical not in matched:
                        matched[canonical] = col_i
            if "start_date" in matched and "name" in matched:
                header_row_idx = row_i
                col_map = matched
                break

        # Fallback: assume row 0 is header with positional columns
        if not col_map:
            col_map = {"start_date": 0, "end_date": 1, "name": 2, "location": 3, "category": 4}

        # ── Read data rows ────────────────────────────────────────────────────
        rows = []
        for r in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
            def _get(key):
                idx = col_map.get(key)
                return r[idx] if idx is not None and idx < len(r) else None

            start_raw = _get("start_date")
            end_raw   = _get("end_date")
            name      = _get("name")
            location  = _get("location")
            category  = _get("category")

            if not name or not start_raw:
                continue
            try:
                start_dt = pd.to_datetime(start_raw)
                end_dt   = pd.to_datetime(end_raw) if end_raw else start_dt
            except Exception:
                continue
            duration = max(1, (end_dt - start_dt).days + 1)
            for d in range(duration):
                rows.append({
                    "Date":          start_dt + pd.Timedelta(days=d),
                    "Event":         str(name).strip(),
                    "Category":      str(category).strip() if category else "",
                    "Location":      str(location).strip() if location else "",
                    "Duration_Days": duration,
                    "Start_Date":    start_dt,
                    "End_Date":      end_dt,
                })
        if not rows:
            return pd.DataFrame()
        df = pd.DataFrame(rows)
        df["Date"] = pd.to_datetime(df["Date"])
        return df.sort_values("Date").reset_index(drop=True)
    except Exception as e:
        raise RuntimeError(f"load_lighthouse_events failed: {e}") from e


def _build_ihg_daily_forecast(year_df: pd.DataFrame, srp_pace_df, total_rooms: int) -> pd.DataFrame:
    """
    Enriches the IHG year DataFrame with daily Forecast_Rooms and Forecast_Rev
    by combining Data_Glance daily OTB with DSS monthly projected totals.

    Algorithm per calendar month:

      Past days (date < today):
        Forecast_Rooms = OTB  (actuals — no forecast needed)
        Revenue_OTB    = OTB × ADR  (SA is forward-only so no SA revenue for past)

      Future days (date >= today):
        remaining_rooms = DSS_proj_rooms - sum(past_OTB_in_month)
        Proportional allocation: each future day gets
          Forecast_Rooms = round(remaining_rooms × day_OTB / sum(future_OTB))
        Integer residual (rounding error) added to the highest-OTB future day.
        Each day is capped at total_rooms (physical capacity).
        Forecast_Rev = Forecast_Rooms × DSS_proj_ADR_for_month

    DSS monthly totals come from srp_pace_df (load_ihg_segment output).
    Proportional allocation preserves the DSS monthly total exactly
    (within physical room constraints).

    Adds column: Forecast_Rev  (new — IHG only)
    Updates:     Forecast_Rooms, Revenue_OTB
    """
    if year_df is None or year_df.empty:
        return year_df

    df = year_df.copy()
    df["Date"] = pd.to_datetime(df["Date"])
    today = pd.Timestamp.today().normalize()
    df["Forecast_Rev"] = None   # new column

    # ── Build DSS monthly totals from srp_pace_df ─────────────────────────────
    dss_monthly = {}   # (year, month) → {proj_rooms, proj_rev, proj_adr}

    if srp_pace_df is not None and not srp_pace_df.empty:
        for _, row in srp_pace_df.iterrows():
            dt = pd.Timestamp(row["Date"])
            key = (dt.year, dt.month)
            if key not in dss_monthly:
                dss_monthly[key] = {"proj_rooms": 0.0, "proj_rev": 0.0}
            dss_monthly[key]["proj_rooms"] += float(row.get("Projected_Rooms", 0) or 0)
            dss_monthly[key]["proj_rev"]   += float(row.get("Projected_Rev",   0) or 0)
        for key in dss_monthly:
            pr = dss_monthly[key]["proj_rooms"]
            rv = dss_monthly[key]["proj_rev"]
            dss_monthly[key]["proj_adr"] = rv / pr if pr > 0 else 0.0

    # ── Process each calendar month ───────────────────────────────────────────
    df["_YM"] = df["Date"].apply(lambda d: (d.year, d.month))

    for ym_key, month_df in df.groupby("_YM"):
        past_idx   = month_df[month_df["Date"] <  today].index
        future_idx = month_df[month_df["Date"] >= today].index

        # Past days: Revenue_OTB = OTB × ADR; Forecast_Rooms = OTB
        # BFR (Booked Forward Revenue from Data_Glance col 5) is used as a fallback
        # for Revenue_OTB when ADR is zero/blank (common for IHG past-closed dates
        # where SA forward-only revenue has not backfilled the actuals). Without this,
        # monthly ADR collapses to near-zero because past_rev ≈ $0 but ty_rooms is
        # the full-month Forecast_Rooms total.
        for idx in past_idx:
            otb = float(df.at[idx, "OTB"] or 0)
            adr = float(df.at[idx, "ADR"] or 0)
            if pd.isna(df.at[idx, "Revenue_OTB"]):
                if otb and adr:
                    df.at[idx, "Revenue_OTB"] = otb * adr
                elif "BFR" in df.columns:
                    bfr = df.at[idx, "BFR"]
                    df.at[idx, "Revenue_OTB"] = float(bfr) if (bfr is not None and not pd.isna(bfr) and float(bfr) > 0) else None
                else:
                    df.at[idx, "Revenue_OTB"] = None
            if pd.isna(df.at[idx, "Forecast_Rooms"]):
                df.at[idx, "Forecast_Rooms"] = otb

        # Future days: proportional allocation from DSS monthly target
        if len(future_idx) == 0:
            continue

        dss = dss_monthly.get(ym_key)
        if dss is None or dss["proj_rooms"] == 0:
            continue

        past_otb_sum   = float(df.loc[past_idx,   "OTB"].fillna(0).sum())
        future_otb_vals = df.loc[future_idx, "OTB"].fillna(0).tolist()
        future_otb_sum  = sum(future_otb_vals)
        if future_otb_sum <= 0:
            continue

        remaining_rooms = max(0.0, dss["proj_rooms"] - past_otb_sum)
        proj_adr        = dss["proj_adr"]

        # Iterative proportional allocation — preserves DSS total within room cap.
        # Exception: dates within 3 days of today use OTB as-is (too close to arrival
        # for DSS monthly scaling to be meaningful — actual OTB is the best forecast).
        # Pass 1: assign proportional shares; cap days that exceed total_rooms.
        # Pass 2+: redistribute capped room-nights among remaining uncapped days.
        # Converges in 2-3 iterations. Integer residual added to highest-OTB days.

        # Mark near-term dates (≤ 3 days out) — use OTB directly, exclude from scaling
        near_term_cutoff = today + pd.Timedelta(days=3)
        near_term_indices = [i for i, idx in enumerate(future_idx)
                             if df.at[idx, "Date"] <= near_term_cutoff]
        scale_indices     = [i for i in range(len(future_otb_vals))
                             if i not in near_term_indices]

        # Assign near-term dates OTB + small pickup buffer as forecast.
        # Pickup factors by days out: today=+3%, 1 day=+3%, 2 days=+6%, 3 days=+8%
        NEAR_TERM_PICKUP = {0: 1.16, 1: 1.24, 2: 1.30, 3: 1.40}
        near_term_rooms = 0.0
        for i in near_term_indices:
            idx      = future_idx[i]
            otb      = float(df.at[idx, "OTB"] or 0)
            days_out = (df.at[idx, "Date"] - today).days
            factor   = NEAR_TERM_PICKUP.get(days_out, 1.0)
            fc_rooms = min(round(otb * factor), total_rooms)
            df.at[idx, "Forecast_Rooms"] = float(fc_rooms)
            # Only write Forecast_Rev if proj_adr is meaningful (non-zero).
            # Writing 0.0 instead of None corrupts the monthly ADR blend in the dashboard.
            df.at[idx, "Forecast_Rev"]   = (fc_rooms * proj_adr) if proj_adr > 0 else None
            near_term_rooms += fc_rooms

        # Adjust remaining rooms target for scale_indices only
        remaining_rooms = max(0.0, remaining_rooms - near_term_rooms)
        future_otb_vals_scale = [future_otb_vals[i] for i in scale_indices]
        future_idx_scale      = [future_idx[i]      for i in scale_indices]

        if not future_idx_scale or sum(future_otb_vals_scale) <= 0:
            continue

        alloc       = [0.0] * len(future_otb_vals_scale)
        rem_float   = float(remaining_rooms)
        uncapped    = list(range(len(future_otb_vals_scale)))

        for _ in range(10):   # max iterations (usually converges in 2-3)
            if not uncapped:
                break
            unc_otb_sum = sum(future_otb_vals_scale[i] for i in uncapped)
            if unc_otb_sum <= 0:
                break
            newly_capped = []
            for i in uncapped:
                share    = future_otb_vals_scale[i] / unc_otb_sum
                alloc[i] = rem_float * share
                if alloc[i] >= total_rooms:
                    alloc[i] = float(total_rooms)
                    newly_capped.append(i)
            if not newly_capped:
                break
            rem_float -= len(newly_capped) * float(total_rooms)
            uncapped   = [i for i in uncapped if i not in newly_capped]

        # Integer rounding — distribute residual to highest-OTB uncapped days
        alloc_int = [int(v) for v in alloc]
        residual  = round(remaining_rooms) - sum(alloc_int)
        if residual > 0:
            for i in sorted(uncapped, key=lambda i: future_otb_vals_scale[i], reverse=True):
                if residual <= 0:
                    break
                if alloc_int[i] < total_rooms:
                    alloc_int[i] += 1
                    residual     -= 1

        for i, idx in enumerate(future_idx_scale):
            fc_rooms = min(alloc_int[i], total_rooms)
            df.at[idx, "Forecast_Rooms"] = float(fc_rooms)
            df.at[idx, "Forecast_Rev"]   = (fc_rooms * proj_adr) if (fc_rooms > 0 and proj_adr > 0) else None

    df = df.drop(columns=["_YM"])
    return df

def _load_all_ihg(files: dict, total_rooms: int, hotel_id: str) -> dict:
    """
    Master loader for IHG hotels. Mirrors the structure of load_all() but
    dispatches to IHG-specific loaders and merges the results into the same
    data dict schema that all existing revpar_app tabs expect.

    File roles used:
      year      → Data_Glance.xlsx       (load_ihg_year)
      strategic → StrategicAnalysis.xlsx (load_ihg_strategic)
      srp_pace  → Property_Segment_Data  (load_ihg_segment)
      budget    → budget.xlsx            (load_budget — unchanged)
      str       → str.xlsx              (load_str_glance — unchanged)
      rates     → rates.xlsx            (load_rates_* — unchanged, if present)
      groups    → group_wash.xlsx       (load_groups — unchanged, if present)
      booking   → SRP Activity          (load_srp — unchanged, if present)

    Roles NOT used for IHG (handled via 'strategic' instead):
      pickup1, pickup7 — pickup comes from StrategicAnalysis cols 13/14
    """
    data = {}

    def safe_call(key, fn, *args):
        try:
            result = fn(*args)
            data[key] = result
        except Exception as e:
            data[key]            = None
            data[f"{key}_error"] = str(e)

    # ── 1. Base year data from Data_Glance ──────────────────────────────────
    if files.get("year"):
        safe_call("_ihg_year_raw", load_ihg_year, files["year"])
    else:
        data["_ihg_year_raw"] = None

    # ── 2. Strategic Analysis — year supplement + rates + pickup ────────────
    if files.get("strategic"):
        safe_call("_ihg_strategic", load_ihg_strategic, files["strategic"], total_rooms)
    else:
        data["_ihg_strategic"] = None

    # ── 3. Merge year df: SA overrides Revenue_OTB and future Forecast_Rooms ─
    year_df = data.get("_ihg_year_raw")
    sa      = data.get("_ihg_strategic")

    if year_df is not None and not year_df.empty and sa is not None:
        supp = sa.get("year_supplement")
        if supp is not None and not supp.empty:
            supp = supp[["Date", "Forecast_Rooms", "Forecast_Current",
                          "Revenue_OTB", "Our_Rate"]].copy()
            supp["Date"] = pd.to_datetime(supp["Date"])
            year_df["Date"] = pd.to_datetime(year_df["Date"])

            # Merge on Date — SA values override Data_Glance for matching dates
            merged = year_df.merge(supp, on="Date", how="left", suffixes=("", "_sa"))

            # Revenue_OTB: always use SA value where available
            merged["Revenue_OTB"] = merged["Revenue_OTB_sa"].combine_first(merged["Revenue_OTB"])

            # Forecast_Rooms: use SA for future dates (SA is forward-looking)
            today = pd.Timestamp.today().normalize()
            future_mask = merged["Date"] >= today
            merged.loc[future_mask, "Forecast_Rooms"] = \
                merged.loc[future_mask, "Forecast_Rooms_sa"].combine_first(
                    merged.loc[future_mask, "Forecast_Rooms"])
            merged.loc[future_mask, "Forecast_Current"] = \
                merged.loc[future_mask, "Forecast_Current_sa"].combine_first(
                    merged.loc[future_mask, "Forecast_Current"])

            # Drop helper columns
            drop_cols = [c for c in merged.columns if c.endswith("_sa")]
            year_df   = merged.drop(columns=drop_cols)

    data["year"] = year_df

    # ── 5. Tonight card ──────────────────────────────────────────────────────
    if data.get("year") is not None:
        try:
            data["tonight"] = get_tonight_otb(data["year"], total_rooms)
        except Exception as e:
            data["tonight"]       = None
            data["tonight_error"] = str(e)
    else:
        data["tonight"] = None

    # ── 6. Build rates data from Rates.xlsx (primary) or Strategic Analysis (fallback) ──
    _rates_file   = files.get("rates")
    _rates_loaded = False

    if _rates_file:
        try:
            data["rates_overview"] = load_rates_overview(_rates_file, hotel_id=hotel_id)
        except Exception as e:
            data["rates_overview"]       = None
            data["rates_overview_error"] = str(e)
        try:
            data["rates_comp"] = load_rates_comp(_rates_file, hotel_id=hotel_id)
        except Exception as e:
            data["rates_comp"]       = None
            data["rates_comp_error"] = str(e)
        try:
            data["rates_vs7"] = load_rates_changes(_rates_file, vs_sheet="vs. 7 days ago", hotel_id=hotel_id)
            _rates_loaded = True
        except Exception:
            try:
                data["rates_vs7"] = load_rates_changes(_rates_file, vs_sheet="vs. Yesterday", hotel_id=hotel_id)
                _rates_loaded = True
            except Exception as e:
                data["rates_vs7"]       = None
                data["rates_vs7_error"] = str(e)

    # If rates_vs7 still None but rates_comp loaded, build synthetic vs7 with no change deltas
    # This ensures Snapshot comp table renders even when no vs7 sheet exists in the file
    if data.get("rates_vs7") is None and data.get("rates_comp") is not None:
        try:
            rc = data["rates_comp"]
            hotels_in_comp = [h for h in rc["Hotel"].unique() if "(us)" not in str(h).lower()]
            our_ov = data.get("rates_overview")
            # Build wide pivot: Date + one __rate col per hotel
            pivot = rc.pivot_table(index="Date", columns="Hotel", values="Rate", aggfunc="first").reset_index()
            pivot.columns.name = None
            rename_map = {h: f"{h}__rate" for h in hotels_in_comp if h in pivot.columns}
            pivot = pivot.rename(columns=rename_map)
            # Add Our_Rate from rates_overview
            if our_ov is not None:
                ov_map = our_ov.set_index("Date")["Our_Rate"].to_dict()
                pivot["Our_Rate (Us)__rate"] = pivot["Date"].map(ov_map)
                pivot["Our_Rate (Us)__change"] = None
            # Add __change = None for all comp hotels
            for h in hotels_in_comp:
                pivot[f"{h}__change"] = None
            data["rates_vs7"] = pivot
        except Exception:
            data["rates_vs7"] = None

    # SA fallback: build rates from Strategic Analysis if Rates.xlsx not available
    if not _rates_loaded and sa is not None:
        rates_df = sa.get("rates")
        if rates_df is not None and not rates_df.empty:
            comp_cols = [c for c in rates_df.columns if c not in ("Date", "Our_Rate")]
            if data.get("rates_overview") is None:
                try:
                    rates_df_num = rates_df[comp_cols].apply(pd.to_numeric, errors="coerce")
                    data["rates_overview"] = pd.DataFrame({
                        "Date":        rates_df["Date"],
                        "Our_Rate":    pd.to_numeric(rates_df["Our_Rate"], errors="coerce"),
                        "Median_Comp": rates_df_num.median(axis=1),
                    })
                except Exception as e:
                    data["rates_overview"]       = None
                    data["rates_overview_error"] = str(e)
            if data.get("rates_vs7") is None:
                try:
                    vs7 = rates_df.copy()
                    rename_map = {col: f"{col}__rate" for col in comp_cols}
                    rename_map["Our_Rate"] = "Our_Rate__rate"
                    vs7 = vs7.rename(columns=rename_map)
                    for col in comp_cols:
                        vs7[f"{col}__change"] = None
                    data["rates_vs7"] = vs7
                except Exception as e:
                    data["rates_vs7"]       = None
                    data["rates_vs7_error"] = str(e)
            if data.get("rates_comp") is None:
                try:
                    comp_long = rates_df.melt(id_vars=["Date"], value_vars=comp_cols,
                                              var_name="Hotel", value_name="Rate")
                    comp_long["Rate_Numeric"] = pd.to_numeric(comp_long["Rate"], errors="coerce")
                    comp_long = comp_long.dropna(subset=["Rate_Numeric"])
                    comp_long = comp_long.sort_values(["Date", "Hotel"]).reset_index(drop=True)
                    data["rates_comp"] = comp_long
                except Exception:
                    data["rates_comp"] = None
        else:
            if data.get("rates_overview") is None: data["rates_overview"] = None
            if data.get("rates_vs7")      is None: data["rates_vs7"]      = None
            if data.get("rates_comp")     is None: data["rates_comp"]     = None

    # ── 7. Pickup from Strategic Analysis ────────────────────────────────
    if sa is not None:
        pickup_df = sa.get("pickup")
        if pickup_df is not None and not pickup_df.empty:
            # Build pickup1 and pickup7 DataFrames matching load_pickup() schema
            # The Snapshot tab reads data["pickup1"] and data["pickup7"]
            def _build_pickup(col_name):
                df = pickup_df[["Date", col_name]].copy()
                df = df.rename(columns={col_name: "Pickup"})
                # Carry OTB, Forecast_Rooms, and Forecast_Current from year_df
                # Forecast_Current is needed by the Dashboard 30-Day chart
                if data.get("year") is not None:
                    yr = data["year"]
                    merge_cols = ["Date", "OTB", "Forecast_Rooms"]
                    if "Forecast_Current" in yr.columns:
                        merge_cols.append("Forecast_Current")
                    df = df.merge(yr[merge_cols], on="Date", how="left")
                return df

            data["pickup1"] = _build_pickup("Pickup1")
            data["pickup7"] = _build_pickup("Pickup7")
        else:
            data["pickup1"] = None
            data["pickup7"] = None
    else:
        data["rates_overview"] = None
        data["rates_vs7"]      = None
        data["rates_comp"]     = None
        data["pickup1"]        = None
        data["pickup7"]        = None

    # ── 8. Segment data (SRP Pace equivalent) ───────────────────────────────
    # Try load_ihg_segment first (old Property_Segment_Data format: 29 cols, sno/Year_Month/Market Segment)
    # If that returns empty, try load_ihg_dss_property (new pivoted format: Measure Name/Year_Month/sno/Value)
    if files.get("srp_pace"):
        try:
            _srp = load_ihg_segment(files["srp_pace"])
            if _srp is None or _srp.empty:
                # Try the pivoted DSS Property Detail Report format
                _srp = load_ihg_dss_property(files["srp_pace"])
            data["srp_pace"] = _srp
        except Exception as e:
            try:
                data["srp_pace"] = load_ihg_dss_property(files["srp_pace"])
            except Exception as e2:
                data["srp_pace"]       = None
                data["srp_pace_error"] = str(e2)
    else:
        data["srp_pace"] = None

    # ── 8b. Enrich year_df with daily forecast from DSS monthly totals ─────
    # Must run AFTER srp_pace is loaded (step 8) since it reads DSS projections
    if data.get("year") is not None:
        try:
            data["year"] = _build_ihg_daily_forecast(
                data["year"],
                data.get("srp_pace"),
                total_rooms,
            )
            # Rebuild tonight card with enriched data
            try:
                data["tonight"] = get_tonight_otb(data["year"], total_rooms)
            except Exception:
                pass
        except Exception as e:
            data["year_forecast_error"] = str(e)

    # ── 9. Budget (IHG-specific loader extracts both budget + LY actuals) ───
    if files.get("budget"):
        try:
            data["budget"] = load_ihg_budget(files["budget"])
        except Exception as e:
            data["budget"]        = None
            data["budget_error"]  = str(e)
    else:
        data["budget"] = None

    # Merge Revenue_LY and LY_Occ_Rooms from IHG budget into year_df
    if data.get("budget") is not None and data.get("year") is not None:
        try:
            bud = data["budget"]
            # Columns available from load_ihg_budget: Revenue_LY, LY_Occ_Rooms, LY_ADR
            ly_cols = [c for c in ["Revenue_LY", "LY_Occ_Rooms", "LY_ADR"] if c in bud.columns]
            if ly_cols:
                bud_ly = bud[["Date"] + ly_cols].copy()
                bud_ly["Date"] = pd.to_datetime(bud_ly["Date"])
                yr = data["year"].copy()
                yr["Date"] = pd.to_datetime(yr["Date"])
                yr = yr.merge(bud_ly, on="Date", how="left", suffixes=("", "_bud"))
                for col in ly_cols:
                    bud_col = f"{col}_bud"
                    if bud_col in yr.columns:
                        yr[col] = yr[bud_col].combine_first(yr.get(col, pd.Series(dtype=float)))
                        yr = yr.drop(columns=[bud_col])
                    elif col not in yr.columns:
                        pass  # column wasn't in year_df and wasn't in budget either
                data["year"] = yr
        except Exception as e:
            data["budget_ly_merge_error"] = str(e)

    # ── 10. STR (unchanged — same format) ────────────────────────────────────
    if files.get("str"):
        safe_call("str", load_str_glance, files["str"])
        try:
            data["str_ranks"] = load_str_ranks(files["str"])
        except Exception:
            data["str_ranks"] = None
        try:
            data["str_seg"] = load_str_segmentation(files["str"])
        except Exception:
            data["str_seg"] = None
    else:
        data["str"]       = None
        data["str_ranks"] = None
        data["str_seg"]   = None

    # ── 11. Groups — IHG uses Group Summary by Day export format ─────────────
    if files.get("groups"):
        safe_call("groups", load_ihg_groups, files["groups"])
    else:
        data["groups"] = None

    # ── 12. Booking / SRP Activity (unchanged — if file present) ─────────────
    if files.get("booking"):
        safe_call("booking", load_srp, files["booking"])
    else:
        data["booking"] = None

    # ── 12b. IHG Corp Segments — corporate account + segment production ─────────
    if files.get("corp_segments"):
        try:
            data["corp_segments"] = load_ihg_corp_segments(files["corp_segments"])
        except Exception as e:
            data["corp_segments"]       = None
            data["corp_segments_error"] = str(e)
    else:
        data["corp_segments"] = None

    # ── 13. Lighthouse Events (IHG only) ─────────────────────────────────────
    if files.get("lighthouse_events"):
        try:
            data["lighthouse_events"] = load_lighthouse_events(files["lighthouse_events"])
        except Exception as e:
            data["lighthouse_events"]       = None
            data["lighthouse_events_error"] = str(e)
    else:
        data["lighthouse_events"] = None

    # ── 14. Flag this data dict as IHG-sourced for tab-level branching ────────
    data["_ihg_hotel"] = True

    return data


# ═══════════════════════════════════════════════════════════════════════════════
# MASTER LOADER  —  Load all files for one hotel at once
# ═══════════════════════════════════════════════════════════════════════════════

def load_all(files: dict, total_rooms: int = 112, hotel_id: str = "hampton_lino_lakes") -> dict:
    """
    Load all available source files for a hotel.
    files: output of hotel_config.detect_files()
    Returns dict keyed by role with DataFrames (or None if file missing).
    IHG hotels are dispatched to _load_all_ihg() automatically.
    """
    # IHG hotels use different source files — dispatch to dedicated loader
    if hotel_id in IHG_HOTELS:
        return _load_all_ihg(files, total_rooms, hotel_id)

    data = {}

    def safe(role, loader, *args):
        path = files.get(role)
        if path:
            try:
                data[role] = loader(path, *args)
            except Exception as e:
                data[role] = None
                data[f"{role}_error"] = str(e)
        else:
            data[role] = None

    safe("year",     load_year)
    safe("budget",   load_budget)
    safe("pickup1",  load_pickup)
    safe("pickup7",  load_pickup)
    safe("groups",   load_groups)
    safe("booking",  load_srp)
    safe("srp_pace", load_srp_pace)
    safe("str",      load_str_glance)

    # STR rank summaries (Current Week / Run 28 / MTD) from Occ, ADR, RevPAR sheets
    if files.get("str"):
        try:
            data["str_ranks"] = load_str_ranks(files["str"])
        except Exception:
            data["str_ranks"] = None
    else:
        data["str_ranks"] = None

    # STR Segmentation Glance (Transient / Group / Contract / Total)
    if files.get("str"):
        try:
            data["str_seg"] = load_str_segmentation(files["str"])
        except Exception:
            data["str_seg"] = None
    else:
        data["str_seg"] = None

    # Rates needs special handling (multiple sub-loaders)
    if files.get("rates"):
        try:
            data["rates_overview"] = load_rates_overview(files["rates"], hotel_id=hotel_id)
        except Exception as e:
            data["rates_overview"] = None
            data["rates_overview_error"] = str(e)
        try:
            data["rates_comp"] = load_rates_comp(files["rates"], hotel_id=hotel_id)
        except Exception as e:
            data["rates_comp"] = None
            data["rates_comp_error"] = str(e)
        try:
            data["rates_vs7"] = load_rates_changes(files["rates"], vs_sheet="vs. 7 days ago", hotel_id=hotel_id)
        except Exception as e:
            data["rates_vs7"] = None
            data["rates_vs7_error"] = str(e)
    else:
        data["rates_overview"] = None
        data["rates_comp"]     = None
        data["rates_vs7"]      = None

    # Tonight card
    if data.get("year") is not None:
        data["tonight"] = get_tonight_otb(data["year"], total_rooms)
    else:
        data["tonight"] = None

    return data
